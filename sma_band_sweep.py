"""Reusable strategy-sweep engine + the SMA-band reversion strategy.

THE SWEPT STRATEGY ("SMA band", S&P 500):
  - Upper band = SMA(window) * (1 + upper_pct),  Lower band = SMA(window) * (1 - lower_pct)
  - ENTRY: price crosses the LOWER band from below (rises up through it)  -> go to `leverage` (1x/2x/3x)
  - EXIT : price crosses the UPPER band from the top (falls down through it) -> go to cash (0x)
  - Hold the current state between crosses.

ENGINE DESIGN (so you can test other strategies):
  A "strategy" is any function  signal_fn(prices, **params) -> pd.Series of daily leverage
  (0 = cash, 1/2/3 = long Nx). `run_backtest` turns a leverage series into an equity curve +
  metrics; `evaluate` wraps one parameter set into a result row; `sweep` runs a strategy across a
  parameter grid. To test a different idea, just write another signal_fn with the same shape and
  hand it to `sweep` (see STRATEGIES registry + sma_band_breakout for a second example).

ASSUMPTIONS: 1-day signal lag, 0.10% cost on leverage turnover, daily-rebalanced synthetic leverage
(2x/3x compound daily and pay funding at the T-bill rate; cash earns the T-bill rate). Full S&P 500
history from yfinance (^GSPC, 1950+).

OUTPUT: Results/SMA band.xlsx (Results + Best sheets + embedded charts) and PNG charts in Results/.

Run:  python sma_band_sweep.py
"""

from __future__ import annotations

import itertools
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from backtest_spx_distance_scale import download_spx_panel
from core.metrics import comprehensive_stats

ROOT = Path(__file__).resolve().parent
RESULTS = ROOT / "Results"
OUT_XLSX = RESULTS / "SMA band.xlsx"
CHART_DIR = RESULTS / "sma_band_charts"

TRADING_DAYS = 252
TRADING_COST = 0.001   # 0.10% of turnover on each leverage change
SIGNAL_LAG = 1         # act on yesterday's signal


# ============================ reusable engine ============================
def run_backtest(prices: pd.DataFrame, leverage: pd.Series, *,
                 trading_cost: float = TRADING_COST, lag: int = SIGNAL_LAG, init: float = 100.0):
    """Turn a daily leverage series into (equity, daily_returns, stats).

    Unified leverage return:  ret = lev*asset_return + (1-lev)*cash_return
    (lev=0 -> earns T-bills; lev>1 -> borrows (lev-1) at the T-bill rate). Turnover cost is charged
    on every change in leverage. Daily rebalanced, so 2x/3x carry the usual leverage compounding drag.
    """
    close = prices["spx_close"].astype(float)
    ret = close.pct_change().fillna(0.0)
    cash_daily = (prices["tbill_rate"].astype(float) / TRADING_DAYS) if "tbill_rate" in prices else 0.0
    lev = leverage.reindex(close.index).shift(lag).fillna(0.0)
    strat_ret = lev * ret + (1.0 - lev) * cash_daily
    turnover = lev.diff().abs().fillna(lev.abs())
    strat_ret = strat_ret - turnover * trading_cost
    equity = init * (1.0 + strat_ret).cumprod()
    rf = float(prices["tbill_rate"].mean()) if "tbill_rate" in prices else 0.0
    stats = comprehensive_stats(equity, strat_ret, risk_free=rf)
    return equity, strat_ret, stats


def evaluate(prices: pd.DataFrame, signal_fn, params: dict) -> dict:
    """Run one parameter set and return a flat metrics row."""
    lev = signal_fn(prices, **params)
    equity, _, st = run_backtest(prices, lev)
    changes = int((lev.diff().fillna(0) != 0).sum())
    years = max((prices.index[-1] - prices.index[0]).days / 365.25, 1e-9)
    return {
        **params,
        "CAGR_pct": _f(st["cagr"] * 100),
        "MaxDD_pct": _f(st["max_drawdown"] * 100),
        "Vol_pct": _f(st["volatility"] * 100),
        "Sharpe": _f(st["sharpe"]),
        "Calmar": _f(st.get("calmar")),
        "End_Value": _f(float(equity.iloc[-1])),
        "Pct_Time_In": _f(float((lev > 0).mean() * 100)),
        "Trades_Per_Year": _f(changes / years),
    }


def sweep(prices: pd.DataFrame, signal_fn, grid: dict, progress: bool = True) -> pd.DataFrame:
    """Run `signal_fn` over the cartesian product of `grid` (param name -> list of values)."""
    keys = list(grid)
    combos = list(itertools.product(*[grid[k] for k in keys]))
    rows = []
    for i, combo in enumerate(combos):
        if progress and i % 100 == 0:
            print(f"  {i}/{len(combos)} combos", flush=True)
        rows.append(evaluate(prices, signal_fn, dict(zip(keys, combo))))
    return pd.DataFrame(rows)


def _f(x):
    """NaN/Inf -> None, else float (keeps Excel/JSON valid)."""
    if x is None:
        return None
    try:
        f = float(x)
    except (TypeError, ValueError):
        return None
    return f if np.isfinite(f) else None


# ============================ strategies (pluggable) ============================
def sma_band_reversion(prices: pd.DataFrame, window: int, upper_pct: float, lower_pct: float,
                       leverage: float) -> pd.Series:
    """ENTER on an upward cross of the lower band; EXIT to cash on a downward cross of the upper band."""
    close = prices["spx_close"].astype(float)
    sma = close.rolling(window, min_periods=window).mean()
    upper = sma * (1.0 + upper_pct)
    lower = sma * (1.0 - lower_pct)
    prev_close = close.shift(1)
    entry = (prev_close < lower.shift(1)) & (close >= lower)   # crossed up through the lower band
    exit_ = (prev_close > upper.shift(1)) & (close <= upper)   # crossed down through the upper band
    sig = pd.Series(np.nan, index=close.index)
    sig[entry] = float(leverage)
    sig[exit_] = 0.0
    return sig.ffill().fillna(0.0)


def sma_band_breakout(prices: pd.DataFrame, window: int, upper_pct: float, lower_pct: float,
                      leverage: float) -> pd.Series:
    """Example 2nd strategy (trend version): ENTER on an upward cross of the UPPER band, EXIT on a
    downward cross of the LOWER band. Same signature -> works with sweep() unchanged."""
    close = prices["spx_close"].astype(float)
    sma = close.rolling(window, min_periods=window).mean()
    upper = sma * (1.0 + upper_pct)
    lower = sma * (1.0 - lower_pct)
    prev_close = close.shift(1)
    entry = (prev_close < upper.shift(1)) & (close >= upper)
    exit_ = (prev_close > lower.shift(1)) & (close <= lower)
    sig = pd.Series(np.nan, index=close.index)
    sig[entry] = float(leverage)
    sig[exit_] = 0.0
    return sig.ffill().fillna(0.0)


STRATEGIES = {
    "sma_band_reversion": sma_band_reversion,
    "sma_band_breakout": sma_band_breakout,
}


# ============================ sweep grid ============================
SMA_WINDOWS = [20, 50, 100, 200]
UPPER_BANDS = [0.0, 0.01, 0.02, 0.03, 0.05, 0.075, 0.10]
LOWER_BANDS = [0.0, 0.01, 0.02, 0.03, 0.05, 0.075, 0.10]
LEVERAGES = [1.0, 2.0, 3.0]
CHART_LEVERAGE = 2.0   # leverage used for the band-sensitivity charts (all leverages are in the table)


# ============================ charts ============================
def make_charts(df: pd.DataFrame) -> list[Path]:
    CHART_DIR.mkdir(parents=True, exist_ok=True)
    paths = []

    # 1) CAGR vs SMA window — best (max-CAGR) band combo per (window, leverage)
    fig, ax = plt.subplots(figsize=(7.2, 4.4))
    for lev in LEVERAGES:
        sub = df[df["leverage"] == lev]
        best = sub.groupby("window")["CAGR_pct"].max().reindex(SMA_WINDOWS)
        ax.plot(SMA_WINDOWS, best.values, marker="o", label=f"{lev:.0f}x")
    ax.set_xlabel("SMA window (days)"); ax.set_ylabel("Best CAGR %")
    ax.set_title("CAGR vs SMA window (best band combo per window)")
    ax.legend(title="Leverage"); ax.grid(alpha=.3); fig.tight_layout()
    p1 = CHART_DIR / "cagr_vs_sma_window.png"; fig.savefig(p1, dpi=130); plt.close(fig); paths.append(p1)

    # 2) Heatmaps CAGR(upper, lower) per SMA window at CHART_LEVERAGE
    sub = df[df["leverage"] == CHART_LEVERAGE]
    vmin = sub["CAGR_pct"].min(); vmax = sub["CAGR_pct"].max()
    fig, axes = plt.subplots(2, 2, figsize=(11, 9.4))
    for k, (ax, w) in enumerate(zip(axes.ravel(), SMA_WINDOWS)):
        piv = sub[sub["window"] == w].pivot(index="lower_pct", columns="upper_pct", values="CAGR_pct")
        piv = piv.reindex(index=LOWER_BANDS, columns=UPPER_BANDS)
        im = ax.imshow(piv.values, origin="lower", aspect="auto", cmap="viridis", vmin=vmin, vmax=vmax)
        ax.set_xticks(range(len(UPPER_BANDS))); ax.set_xticklabels([f"{u*100:g}" for u in UPPER_BANDS])
        ax.set_yticks(range(len(LOWER_BANDS))); ax.set_yticklabels([f"{l*100:g}" for l in LOWER_BANDS])
        if k >= 2: ax.set_xlabel("Upper band %")
        if k % 2 == 0: ax.set_ylabel("Lower band %")
        ax.set_title(f"SMA {w}d  ({CHART_LEVERAGE:.0f}x)")
        for i in range(len(LOWER_BANDS)):
            for j in range(len(UPPER_BANDS)):
                v = piv.values[i, j]
                if np.isfinite(v):
                    ax.text(j, i, f"{v:.0f}", ha="center", va="center", color="white", fontsize=7)
    fig.subplots_adjust(hspace=0.22, wspace=0.12, top=0.91, bottom=0.07, left=0.07, right=0.9)
    fig.colorbar(im, ax=axes, shrink=.7, label="CAGR %")
    fig.suptitle(f"CAGR by upper × lower band, per SMA window ({CHART_LEVERAGE:.0f}x leverage)", y=.965)
    p2 = CHART_DIR / "cagr_band_heatmaps.png"; fig.savefig(p2, dpi=130); plt.close(fig); paths.append(p2)

    # 3) Marginal effect of each band (max CAGR over the other band) at CHART_LEVERAGE
    fig, (axu, axl) = plt.subplots(1, 2, figsize=(11, 4.4))
    for w in SMA_WINDOWS:
        s = sub[sub["window"] == w]
        mu = s.groupby("upper_pct")["CAGR_pct"].max().reindex(UPPER_BANDS)
        axu.plot([u * 100 for u in UPPER_BANDS], mu.values, marker="o", label=f"SMA {w}d")
        ml = s.groupby("lower_pct")["CAGR_pct"].max().reindex(LOWER_BANDS)
        axl.plot([l * 100 for l in LOWER_BANDS], ml.values, marker="o", label=f"SMA {w}d")
    axu.set_xlabel("Upper band %"); axu.set_ylabel("Best CAGR %"); axu.set_title("CAGR vs upper band"); axu.grid(alpha=.3); axu.legend()
    axl.set_xlabel("Lower band %"); axl.set_ylabel("Best CAGR %"); axl.set_title("CAGR vs lower band"); axl.grid(alpha=.3); axl.legend()
    fig.suptitle(f"Marginal band sensitivity ({CHART_LEVERAGE:.0f}x; best over the other band)")
    fig.tight_layout(rect=(0, 0, 1, .95))
    p3 = CHART_DIR / "cagr_vs_bands.png"; fig.savefig(p3, dpi=130); plt.close(fig); paths.append(p3)
    return paths


# ============================ excel ============================
def write_excel(df: pd.DataFrame, bh: dict, chart_paths: list[Path]) -> None:
    RESULTS.mkdir(parents=True, exist_ok=True)
    cols = ["window", "upper_pct", "lower_pct", "leverage", "CAGR_pct", "MaxDD_pct",
            "Vol_pct", "Sharpe", "Calmar", "End_Value", "Pct_Time_In", "Trades_Per_Year"]
    nice = {"window": "SMA Window", "upper_pct": "Upper Band %", "lower_pct": "Lower Band %",
            "leverage": "Leverage", "CAGR_pct": "CAGR %", "MaxDD_pct": "Max DD %", "Vol_pct": "Vol %",
            "Sharpe": "Sharpe", "Calmar": "Calmar", "End_Value": "End $ ($100→)",
            "Pct_Time_In": "% Time Invested", "Trades_Per_Year": "Trades/yr"}
    out = df[cols].copy()
    out["upper_pct"] = out["upper_pct"] * 100
    out["lower_pct"] = out["lower_pct"] * 100
    out = out.rename(columns=nice)
    results = out.sort_values("CAGR %", ascending=False)
    best_each = (df.sort_values("CAGR_pct", ascending=False)
                 .groupby(["window", "leverage"], as_index=False).first()[cols]
                 .sort_values(["leverage", "window"]))
    best_each["upper_pct"] *= 100; best_each["lower_pct"] *= 100
    best_each = best_each.rename(columns=nice)

    bh_row = pd.DataFrame([{
        "SMA Window": "—", "Upper Band %": "—", "Lower Band %": "—", "Leverage": "Buy & Hold 1x",
        "CAGR %": bh["cagr"] * 100, "Max DD %": bh["max_drawdown"] * 100, "Vol %": bh["volatility"] * 100,
        "Sharpe": bh["sharpe"], "Calmar": bh.get("calmar"), "End $ ($100→)": bh["end_value"],
        "% Time Invested": 100.0, "Trades/yr": 0.0,
    }])

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as xw:
        pd.concat([bh_row, results], ignore_index=True).to_excel(xw, sheet_name="Results", index=False)
        pd.concat([bh_row, best_each], ignore_index=True).to_excel(xw, sheet_name="Best per window×lev", index=False)

    wb = openpyxl.load_workbook(OUT_XLSX)
    pct2 = "0.00"
    fmt = {"CAGR %": pct2, "Max DD %": pct2, "Vol %": pct2, "Sharpe": "0.000", "Calmar": "0.00",
           "End $ ($100→)": "#,##0", "% Time Invested": "0.0", "Trades/yr": "0.0",
           "Upper Band %": "0.0", "Lower Band %": "0.0"}
    for sheet in ("Results", "Best per window×lev"):
        ws = wb[sheet]
        hdr = {c.value: c.column for c in ws[1]}
        for col, f in fmt.items():
            if col in hdr:
                letter = get_column_letter(hdr[col])
                for cell in ws[letter][1:]:
                    cell.number_format = f
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="305496")
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(width + 2, 9), 16)

    ws = wb.create_sheet("Charts")
    ws["A1"] = "SMA-band sweep — CAGR sensitivity (S&P 500, full history)"
    ws["A1"].font = Font(bold=True, size=13)
    anchor_row = 3
    for p in chart_paths:
        img = XLImage(str(p))
        ws.add_image(img, f"A{anchor_row}")
        anchor_row += int(img.height / 18) + 3
    wb.save(OUT_XLSX)


# ============================ main ============================
def main() -> int:
    print("Downloading full S&P 500 history (^GSPC + T-bill)...", flush=True)
    prices = download_spx_panel()
    print(f"Loaded {len(prices)} sessions: {prices.index[0].date()} -> {prices.index[-1].date()}", flush=True)

    bh_lev = pd.Series(1.0, index=prices.index)
    bh_eq, _, bh_st = run_backtest(prices, bh_lev)
    bh = {"cagr": bh_st["cagr"], "max_drawdown": bh_st["max_drawdown"], "volatility": bh_st["volatility"],
          "sharpe": bh_st["sharpe"], "calmar": bh_st.get("calmar"), "end_value": float(bh_eq.iloc[-1])}
    print(f"Buy & Hold 1x: CAGR {bh['cagr']*100:.2f}%  MaxDD {bh['max_drawdown']*100:.2f}%  "
          f"Sharpe {bh['sharpe']:.3f}  End ${bh['end_value']:,.0f}", flush=True)

    grid = {"window": SMA_WINDOWS, "upper_pct": UPPER_BANDS, "lower_pct": LOWER_BANDS, "leverage": LEVERAGES}
    n = len(SMA_WINDOWS) * len(UPPER_BANDS) * len(LOWER_BANDS) * len(LEVERAGES)
    print(f"Sweeping {n} combinations of SMA-band reversion...", flush=True)
    df = sweep(prices, sma_band_reversion, grid)

    top = df.sort_values("CAGR_pct", ascending=False).head(8)
    print("\nTop 8 by CAGR:")
    for _, r in top.iterrows():
        print(f"  SMA{int(r['window']):>3}d  +{r['upper_pct']*100:>4.1f}/-{r['lower_pct']*100:>4.1f}%  "
              f"{r['leverage']:.0f}x | CAGR {r['CAGR_pct']:.2f}%  DD {r['MaxDD_pct']:.2f}%  "
              f"Sharpe {r['Sharpe']:.3f}  Calmar {r['Calmar']:.2f}  End ${r['End_Value']:,.0f}")

    print("\nBuilding charts...", flush=True)
    chart_paths = make_charts(df)
    print("Writing Excel...", flush=True)
    write_excel(df, bh, chart_paths)
    print(f"\nWrote {OUT_XLSX}")
    print(f"Charts: {', '.join(p.name for p in chart_paths)} (in {CHART_DIR})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
