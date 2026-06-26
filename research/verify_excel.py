import sys as _s, pathlib as _p; _s.path.insert(0, str(_p.Path(__file__).resolve().parent.parent))  # repo root importable (moved into research/)
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Results/strategy_results.xlsx')

# Print sheet names
print('=== SHEETS ===')
for s in wb.sheetnames:
    print(f'  {s}')
print()

# Check Summary sheet for SPX and NDX counts
ws = wb['Summary']
print('=== SUMMARY SHEET (all rows, first 5 cols) ===')
for row in range(1, ws.max_row+1):
    vals = []
    for col in range(1, min(6, ws.max_column+1)):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals:
        print(f'  Row {row}: {" | ".join(vals)}')
print()

# Check S&P 500 Water sheet - count rows
ws2 = wb['S&P 500 Water']
print(f'=== S&P 500 Water sheet: {ws2.max_row} rows, {ws2.max_column} cols ===')
# Print first 10 rows
for row in range(1, min(11, ws2.max_row+1)):
    vals = []
    for col in range(1, min(6, ws2.max_column+1)):
        v = ws2.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals:
        print(f'  Row {row}: {" | ".join(vals)}')
print()

# Check S&P 500 Octane sheet
ws3 = wb['S&P 500 Octane']
print(f'=== S&P 500 Octane sheet: {ws3.max_row} rows, {ws3.max_column} cols ===')
for row in range(1, min(11, ws3.max_row+1)):
    vals = []
    for col in range(1, min(6, ws3.max_column+1)):
        v = ws3.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals:
        print(f'  Row {row}: {" | ".join(vals)}')
print()

# Check Nasdaq 100 Water sheet
ws4 = wb['Nasdaq 100 Water']
print(f'=== Nasdaq 100 Water sheet: {ws4.max_row} rows, {ws4.max_column} cols ===')
for row in range(1, min(11, ws4.max_row+1)):
    vals = []
    for col in range(1, min(6, ws4.max_column+1)):
        v = ws4.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals:
        print(f'  Row {row}: {" | ".join(vals)}')
print()

# Check Nasdaq 100 Octane sheet
ws5 = wb['Nasdaq 100 Octane']
print(f'=== Nasdaq 100 Octane sheet: {ws5.max_row} rows, {ws5.max_column} cols ===')
for row in range(1, min(11, ws5.max_row+1)):
    vals = []
    for col in range(1, min(6, ws5.max_column+1)):
        v = ws5.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals:
        print(f'  Row {row}: {" | ".join(vals)}')