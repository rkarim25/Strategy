"""Rewrite ONLY the 'S&P 500 Water' and 'S&P 500 Octane' tabs in the workbook:
canonical incumbents (build_strategies) + the band-sweep winners, all recomputed
on ONE consistent data vintage via the real run_one_backtest, then classified and
written with the existing build_detail_sheet styler. Deletes the 4 Band-Sweep sheets.

Usage: python regen_spx_tabs.py <base_xlsx> <dest_xlsx>
"""
from __future__ import annotations
import sys
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

import sweep_all_assets_strategies as sw
import build_strategy_results_excel as bx
from core.indicators import sma as sma_ind
import signals as S

BASE = Path(sys.argv[1]); DEST = Path(sys.argv[2])

# ---- 1. load SPX exactly as the master sweep does ----
print("Downloading T-bill/VIX + SPX (sweep loader)...", flush=True)
tbill_g, vix_g = sw._download_tbill_vix()
ad = sw.load_asset_data("spx", tbill_g, vix_g)
prices = ad["prices"]
cfg = sw.ASSETS["spx"]
etp_panel = sw.build_asset_etp_panel(prices, ad["etp_closes"], cfg)
print(f"  {len(prices)} rows {prices.index[0].date()}..{prices.index[-1].date()}  cost={cfg['trading_cost']}", flush=True)

# ---- 2. canonical incumbents + band-sweep winners ----
strategies = sw.build_strategies(prices, "spx")
print(f"  incumbents: {len(strategies)}", flush=True)

close = prices["spx_close"].to_numpy(dtype=float)
sma200 = sma_ind(prices["spx_close"], 200).to_numpy(dtype=float)
idx = prices.index

def accel(N, lev):
    sig = S.variant_b_accel(close, sma200, 0.03, "conv", N, 0.02, 0.03)
    return pd.Series(sig * lev, index=idx)

def band(up, lo, lev):
    sig = S.conv_band(close, sma200, up, lo)
    return pd.Series(sig * lev, index=idx)

winners = [
    ("SMA200 ±3% Band + Accel-Exit N10 1x/cash", accel(10, 1.0), 1.0),
    ("SMA200 ±3% Band + Accel-Exit N20 1x/cash", accel(20, 1.0), 1.0),
    ("SMA200 ±3% Band + Accel-Exit N10 2x/cash", accel(10, 2.0), 2.0),
    ("SMA200 ±3% Band + Accel-Exit N20 2x/cash", accel(20, 2.0), 2.0),
    ("SMA200 +5%/-3% Band 2x/cash", band(0.05, 0.03, 2.0), 2.0),
]
strategies = strategies + winners
print(f"  + winners: {len(winners)}  -> total {len(strategies)}", flush=True)

# ---- 3. run all through the real engine ----
rows = []
for i, (name, lev, lev_max) in enumerate(strategies):
    r = sw.run_one_backtest(prices, etp_panel, name, lev, cfg["trading_cost"])
    rows.append({
        "Asset": "spx", "Strategy": name, "Leverage_Max": lev_max,
        "CAGR_pct": r["cagr"] * 100, "Vol_pct": r["volatility"] * 100,
        "Sharpe": r["sharpe"], "Sortino": r["sortino"], "Calmar": r["calmar"],
        "MaxDD_pct": r["max_drawdown"] * 100, "End_Value": r["end_value"],
        "Start_Date": str(r["start_date"])[:10], "End_Date": str(r["end_date"])[:10],
        "Years": r["years"], "Pct_Cash_Time": r["pct_cash_time"],
        "Trades_Per_Year": r["trades_per_year"], "Total_Trades": r["total_trades"],
        "Avg_Leverage": r["avg_leverage"],
    })
    tag = "  <-- WINNER" if name in {w[0] for w in winners} else ""
    print(f"    [{i+1:2d}/{len(strategies)}] {name:48s} CAGR {r['cagr']*100:6.2f} DD {r['max_drawdown']*100:7.2f} "
          f"Sharpe {r['sharpe']:.3f} Calmar {r['calmar']:.3f}{tag}", flush=True)

df = pd.DataFrame(rows)
df = bx.classify_all(df)   # adds 'Classification'

print("\nWinner classifications:")
for w in winners:
    row = df[df.Strategy == w[0]].iloc[0]
    print(f"  {w[0]:48s} -> {row['Classification']}")
print("\nClass counts:", df["Classification"].value_counts().to_dict())

# ---- 4. rewrite the two SPX sheets in the workbook ----
wb = openpyxl.load_workbook(BASE)
# drop the 4 band-sweep sheets
for s in ["SPX Band-Sweep Water", "SPX Band-Sweep Octane", "SPX Band-Sweep Method", "SPX Band-Sweep Diagrams"]:
    if s in wb.sheetnames:
        del wb[s]
# remember original positions of the two SPX sheets
order = wb.sheetnames[:]
pos_water = order.index("S&P 500 Water")
pos_octane = order.index("S&P 500 Octane")
del wb["S&P 500 Water"]; del wb["S&P 500 Octane"]
ws_w = bx.build_detail_sheet(wb, df, "spx", "Water")
ws_o = bx.build_detail_sheet(wb, df, "spx", "Octane")
# move them back to original positions (water then octane, adjacent)
wb._sheets.remove(ws_w); wb._sheets.remove(ws_o)
wb._sheets.insert(pos_water, ws_w)
wb._sheets.insert(pos_octane, ws_o)

wb.save(DEST)
print(f"\nWROTE {DEST}")
print("Final sheet order:", wb.sheetnames)
