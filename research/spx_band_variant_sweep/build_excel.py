"""Append S&P 500 band-sweep results (Water/Octane winners only), theoretical
diagrams, and a methodology sheet to a COPY of strategy_results.xlsx.

Usage: python build_excel.py <out_dir> <base_xlsx> <dest_xlsx>
"""
from __future__ import annotations
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

OUT = Path(sys.argv[1]); BASE = Path(sys.argv[2]); DEST = Path(sys.argv[3])

# ---- styles (mirror build_strategy_results_excel.py) ----
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BH_FILL = PatternFill('solid', fgColor='BDD7EE'); BH_FONT = Font(name='Calibri', size=11, bold=True)
WIN_FILL = PatternFill('solid', fgColor='C6EFCE')
INC_FILL = PatternFill('solid', fgColor='FFF2CC'); INC_FONT = Font(name='Calibri', size=11, bold=True, italic=True)
NORMAL = Font(name='Calibri', size=11)
TB = Border(*[Side(style='thin', color='D0D0D0')] * 4)
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='1F4E79')
SUB_FONT = Font(name='Calibri', size=9, italic=True, color='808080')
SEC_FONT = Font(name='Calibri', size=11, bold=True, color='375623')

DETAIL_COLS = ['Strategy', 'Leverage_Max', 'CAGR_pct', 'Vol_pct', 'Sharpe', 'Sortino', 'Calmar',
               'MaxDD_pct', 'End_Value', 'Start_Date', 'End_Date', 'Years', 'Pct_Cash_Time',
               'Trades_Per_Year', 'Total_Trades', 'Avg_Leverage']
NUMFMT = {'CAGR_pct': '0.00', 'Vol_pct': '0.00', 'MaxDD_pct': '0.00', 'Pct_Cash_Time': '0.00',
          'Sharpe': '0.000', 'Sortino': '0.000', 'Calmar': '0.000', 'Avg_Leverage': '0.000',
          'Trades_Per_Year': '0.000', 'End_Value': '$#,##0', 'Total_Trades': '0', 'Years': '0.0',
          'Leverage_Max': '0.0'}

df = pd.read_csv(OUT / 'spx_band_variants_all.csv')
base = pd.read_csv(OUT / 'spx_baselines.csv')
START, END, YEARS = base.iloc[0]['Start_Date'] if 'Start_Date' in base else None, None, None
# baselines rows: BH1,BH2,BH3, inc_water, inc_octane
bh1 = base.iloc[0]; bh2 = base.iloc[1]; bh3 = base.iloc[2]
inc_water = base.iloc[3]; inc_oct = base.iloc[4]

META = pd.read_csv(OUT / 'meta.csv').iloc[0]
SDATE, EDATE, YRS = META['start'], META['end'], float(META['years'])


def row_vals(r, lev_max):
    return {
        'Strategy': r['Strategy'], 'Leverage_Max': lev_max, 'CAGR_pct': r['CAGR_pct'],
        'Vol_pct': r['Vol_pct'], 'Sharpe': r['Sharpe'], 'Sortino': r['Sortino'],
        'Calmar': r['Calmar'], 'MaxDD_pct': r['MaxDD_pct'], 'End_Value': r['End_Value'],
        'Start_Date': SDATE, 'End_Date': EDATE, 'Years': YRS, 'Pct_Cash_Time': r['Pct_Cash_Time'],
        'Trades_Per_Year': r['Trades_Per_Year'], 'Total_Trades': r['Total_Trades'],
        'Avg_Leverage': r['Avg_Leverage'],
    }


def write_header(ws, row):
    for ci, cn in enumerate(DETAIL_COLS, 1):
        c = ws.cell(row=row, column=ci, value=cn)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = TB
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return row + 1


def write_data_row(ws, row, vals, fill, font):
    for ci, cn in enumerate(DETAIL_COLS, 1):
        v = vals.get(cn)
        c = ws.cell(row=row, column=ci, value=v if pd.notna(v) else '')
        c.font = font; c.fill = fill; c.border = TB
        if cn in NUMFMT and pd.notna(v) and v != '':
            c.number_format = NUMFMT[cn]
    return row + 1


def build_sheet(wb, title, klass, winners, inc_rows, intro):
    ws = wb.create_sheet(title)
    r = 1
    c = ws.cell(row=r, column=1, value=f'S&P 500 — Band-Sweep {klass} (NEW: beats incumbent {klass})')
    c.font = TITLE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(DETAIL_COLS)); r += 1
    c = ws.cell(row=r, column=1, value=intro)
    c.font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(DETAIL_COLS)); r += 1
    c = ws.cell(row=r, column=1, value=f'Full history {SDATE} … {EDATE} ({YRS:.1f}y) · 0.10% cost · 1-day signal lag · $10/yr inflow · real SPY/SSO/UPRO ETP returns. Only strategies that BEAT the incumbent are listed.')
    c.font = SUB_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(DETAIL_COLS)); r += 2

    # benchmarks + incumbents
    c = ws.cell(row=r, column=1, value='Benchmarks & current incumbent (reference):'); c.font = BH_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(DETAIL_COLS)); r += 1
    r = write_header(ws, r)
    for ref in (bh1, bh2, bh3):
        r = write_data_row(ws, r, row_vals(ref, ref['Leverage']), BH_FILL, BH_FONT)
    for ref in inc_rows:
        r = write_data_row(ws, r, row_vals(ref, ref['Leverage']), INC_FILL, INC_FONT)
    r += 1

    c = ws.cell(row=r, column=1, value=f'NEW {klass} strategies that beat the incumbent ({len(winners)}):'); c.font = SEC_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(DETAIL_COLS)); r += 1
    r = write_header(ws, r)
    for _, w in winners.iterrows():
        r = write_data_row(ws, r, row_vals(w, w['Leverage']), WIN_FILL, NORMAL)

    widths = [46, 11, 9, 8, 8, 8, 8, 9, 13, 11, 11, 7, 11, 13, 12, 12]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    ws.freeze_panes = 'A6'
    return ws


def main():
    wb = openpyxl.load_workbook(BASE)

    # ----- derive winners: STRICT Pareto-vs-incumbent (no degenerate deep-DD leaks) -----
    iwS, iwC, iwD, iwG, iwO, iwV = (inc_water['Sharpe'], inc_water['Calmar'], inc_water['MaxDD_pct'],
                                    inc_water['CAGR_pct'], inc_water['Sortino'], inc_water['Vol_pct'])
    water = df[df.Class == 'Water'].copy()
    better_water = water[(water.MaxDD_pct >= iwD - 0.05) & (water.Sharpe >= iwS) & (water.Calmar >= iwC) &
                         (water.CAGR_pct >= iwG) & (water.Sortino >= iwO) & (water.Vol_pct <= iwV) &
                         ((water.Sharpe > iwS) | (water.Calmar > iwC) | (water.CAGR_pct > iwG) | (water.MaxDD_pct > iwD))]
    better_water = better_water.sort_values('Sharpe', ascending=False)

    ioC, ioD, ioG = inc_oct['Calmar'], inc_oct['MaxDD_pct'], inc_oct['CAGR_pct']
    octane = df[df.Class == 'Octane'].copy()
    better_octane = octane[(octane.Calmar > ioC) & (octane.MaxDD_pct >= ioD) &
                           (octane.CAGR_pct >= ioG - 1.0) & (octane.Trades_Per_Year <= 30.0)]
    better_octane = better_octane.sort_values('Calmar', ascending=False).head(40)
    n_oct3x = len(df[(df.Class == 'Octane') & (df.Leverage == 3)])

    build_sheet(wb, 'SPX Band-Sweep Water', 'Water', better_water, [inc_water],
                "Water = no metric worse than Buy & Hold 1x, strictly better on CAGR or MaxDD. "
                "Incumbent reference = the S&P Water band strategy (SMA200 ±3% 1x/cash).")
    build_sheet(wb, 'SPX Band-Sweep Octane', 'Octane', better_octane, [inc_oct],
                "Octane = CAGR>B&H1x, Calmar>B&H1x, MaxDD≥-45%, ≤30 trades/yr. "
                "Incumbent reference = SMA200 ±3% Band + RSI>20 Exit 2x.")

    # ----- methodology sheet -----
    ws = wb.create_sheet('SPX Band-Sweep Method')
    lines = [
        ('S&P 500 SMA-Band Variant Sweep — Methodology', TITLE_FONT),
        (f'As-of {EDATE}. Full S&P 500 history {SDATE} … {EDATE} ({YRS:.1f}y). ~{len(df):,} strategies screened.', SUB_FONT),
        ('', NORMAL),
        ('Engine (identical to the main Excel sweep, validated to reproduce it to the penny):', SEC_FONT),
        ('  • 0.10% trading cost from mid on each leverage change · 1-day signal lag (no look-ahead).', NORMAL),
        ('  • $100 base + $10/yr absolute inflow · VIX-linked borrow on leverage · real SPY/SSO/UPRO ETP returns (synthetic pre-inception).', NORMAL),
        ('  • Daily-CLOSE model (no intraday): stops and band crosses are evaluated on the daily close.', NORMAL),
        ('', NORMAL),
        ('Common grid: SMA ∈ {20,50,100,200} · leverage ∈ {1x,2x,3x} · bands symmetric {1,2,3,5%} and asymmetric pairs · stop-loss ∈ {none,0.5,1,1.5,2%}.', NORMAL),
        ('', NORMAL),
        ('Variant A — band entry & exit + fixed stop-loss. Two band directions tested:', SEC_FONT),
        ('   conv  = conventional (enter ABOVE +upper band, exit BELOW -lower band) — the incumbent logic.', NORMAL),
        ('   early = your rule: enter rising THROUGH the -lower band, exit falling THROUGH the +upper band (early-in/early-out), with a -lower-band breakdown failsafe.', NORMAL),
        ("Variant A' — same bands, but a TRAILING stop (3/5/8% from the running peak) instead of a fixed stop.", NORMAL),
        ('', NORMAL),
        ('Variant B — band entry, momentum-based exit ("keep jumping bands or leave"). Two readings:', SEC_FONT),
        ('   decay = exit when the premium (close/SMA-1) falls ≥ d from its trailing N-day max (cushion ratchets 3→2→1%). N∈{5,10,20}, d∈{1,2%}.', NORMAL),
        ('   accel = exit if, N days after entry, price has not made a NEW band-high ≥ step above the entry premium. N∈{5,10,20}, step=2%.', NORMAL),
        ('   (optional fixed stop layered on; -lower-band breakdown failsafe.)', NORMAL),
        ('', NORMAL),
        ('Variant C — momentum ENTRY, band EXIT (+ optional stop).', SEC_FONT),
        ('   entry = RSI(14) crosses up through {30,50}, or MACD(12,26,9) bullish crossover.', NORMAL),
        ('   exit  = close falls through the +upper band (or breaks the -lower band).', NORMAL),
        ('', NORMAL),
        ('Classification (vs a FRESHLY-computed Buy & Hold 1x on this exact window):', SEC_FONT),
        (f'   B&H 1x: CAGR {bh1["CAGR_pct"]:.2f}% · Vol {bh1["Vol_pct"]:.2f}% · Sharpe {bh1["Sharpe"]:.3f} · Sortino {bh1["Sortino"]:.3f} · Calmar {bh1["Calmar"]:.3f} · MaxDD {bh1["MaxDD_pct"]:.2f}%.', NORMAL),
        ('   Water  = no metric worse than B&H 1x (Sharpe/Sortino/Calmar/MaxDD/Vol/CAGR) AND strictly better on CAGR or MaxDD.', NORMAL),
        ('   Octane = CAGR>B&H 1x AND Calmar>B&H 1x AND MaxDD≥-45% AND ≤30 trades/yr.', NORMAL),
        ('   Only strategies that ALSO beat the current S&P incumbent (Water: Sharpe/Calmar of SMA200±3% 1x; Octane: Calmar of SMA200±3%+RSI>20 2x) are reported. Failed/duplicate results are omitted.', NORMAL),
        ('', NORMAL),
        (f'Incumbent Water  (SMA200 ±3% 1x/cash):  CAGR {inc_water["CAGR_pct"]:.2f}% · Sharpe {inc_water["Sharpe"]:.3f} · Calmar {inc_water["Calmar"]:.3f} · MaxDD {inc_water["MaxDD_pct"]:.2f}%.', NORMAL),
        (f'Incumbent Octane (SMA200 ±3% + RSI>20 2x): CAGR {inc_oct["CAGR_pct"]:.2f}% · Calmar {inc_oct["Calmar"]:.3f} · MaxDD {inc_oct["MaxDD_pct"]:.2f}% · {inc_oct["Trades_Per_Year"]:.1f} trades/yr.', NORMAL),
        ('', NORMAL),
        ('Headline finding:', SEC_FONT),
        ('   The variant-B "must-accelerate" exit (cut a trade that fails to make a new +2% higher band within 10–20 days of entry) layered on the SMA200 ±3% entry', NORMAL),
        ('   beats BOTH incumbents — at 1x it is a no-regret (Pareto) improvement on Water; at 2x it improves the Octane Calmar/Sharpe/drawdown. This is exactly the', NORMAL),
        ('   "once in a trade with no momentum, don\'t hang around" thesis. The improvement is modest but consistent across N=10 and N=20.', NORMAL),
        (f'   3x leverage tested: {n_oct3x} signals reached Octane-CLASS at 3x but NONE beat the 2x incumbent, and ZERO 3x strategies qualified as Water — the -45% drawdown gate caps 3x (a known structural fact for the S&P).', NORMAL),
    ]
    for i, (txt, fnt) in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=txt); c.font = fnt
    ws.column_dimensions['A'].width = 160

    # ----- diagrams sheet -----
    wd = wb.create_sheet('SPX Band-Sweep Diagrams')
    wd.cell(row=1, column=1, value='Theoretical entry/exit mechanics — what each rule captures (schematic, not backtest output)').font = TITLE_FONT
    anchor_row = 3
    for png in ['diag_bands.png', 'diag_stops.png', 'diag_variantB.png', 'diag_variantC.png']:
        p = OUT / png
        if p.exists():
            img = XLImage(str(p))
            img.width = int(img.width * 0.85); img.height = int(img.height * 0.85)
            wd.add_image(img, f'A{anchor_row}')
            anchor_row += 26
    wd.column_dimensions['A'].width = 20

    wb.save(DEST)
    print(f'WROTE {DEST}  |  Water winners {len(better_water)}  Octane winners {len(better_octane)}')
    print(f'Sheets: {wb.sheetnames[-4:]}')


if __name__ == '__main__':
    main()
