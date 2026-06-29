"""Add the band-sweep winners to the EXISTING 'S&P 500 Water' / 'S&P 500 Octane'
tabs without losing any incumbent: load the 179-strategy source CSV (frozen basis),
append the 5 winners computed on the SAME window (truncated to the CSV end date),
re-classify, rebuild the two tabs with build_detail_sheet, drop the Band-Sweep sheets.

Usage: python add_winners.py <base_xlsx> <dest_xlsx>
"""
from __future__ import annotations
import sys
from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl

import sweep_all_assets_strategies as sw
import build_strategy_results_excel as bx
from core.indicators import sma as sma_ind
import signals as S

REPO = Path(r"C:/Users/Reza Karim/OneDrive/Systematic_Backstester")
BASE = Path(sys.argv[1]); DEST = Path(sys.argv[2])

# ---- 1. incumbents (frozen source of the existing sheet) ----
inc = pd.read_csv(REPO / "output/strategy_results/spx_results.csv")
inc.insert(0, "Asset", "spx")
END = str(inc[inc.Strategy == "Buy & Hold 1x"].iloc[0]["End_Date"])[:10]
print(f"Incumbents: {len(inc)} rows, frozen end {END}", flush=True)

# ---- 2. load SPX (sweep loader) truncated to the SAME end date ----
tbill_g, vix_g = sw._download_tbill_vix()
ad = sw.load_asset_data("spx", tbill_g, vix_g)
prices = ad["prices"]
prices = prices[prices.index <= END].copy()
cfg = sw.ASSETS["spx"]
etp_panel = sw.build_asset_etp_panel(prices, ad["etp_closes"], cfg)
print(f"  loaded+truncated: {len(prices)} rows ..{prices.index[-1].date()}", flush=True)

close = prices["spx_close"].to_numpy(dtype=float)
sma200 = sma_ind(prices["spx_close"], 200).to_numpy(dtype=float)
idx = prices.index

def accel(N, lev):
    return pd.Series(S.variant_b_accel(close, sma200, 0.03, "conv", N, 0.02, 0.03) * lev, index=idx)
def band(up, lo, lev):
    return pd.Series(S.conv_band(close, sma200, up, lo) * lev, index=idx)

winners = [
    ("SMA200 +-3% Band + Accel-Exit N10 1x/cash", accel(10, 1.0), 1.0),
    ("SMA200 +-3% Band + Accel-Exit N20 1x/cash", accel(20, 1.0), 1.0),
    ("SMA200 +-3% Band + Accel-Exit N10 2x", accel(10, 2.0), 2.0),
    ("SMA200 +-3% Band + Accel-Exit N20 2x", accel(20, 2.0), 2.0),
    ("SMA200 +5%/-3% Band 2x", band(0.05, 0.03, 2.0), 2.0),
]

# sanity reference: plain band 1x should reproduce the CSV value (~10.48)
ref = sw.run_one_backtest(prices, etp_panel, "ref", band(0.03, 0.03, 1.0), cfg["trading_cost"])
csv_band = float(inc[inc.Strategy == "SMA200 +-3% Band 1x/cash"].iloc[0]["CAGR_pct"])
print(f"  basis check: plain band 1x CAGR {ref['cagr']*100:.2f} (CSV {csv_band:.2f})", flush=True)

# ---- 3. run winners, map to CSV columns ----
bh1 = sw.run_one_backtest(prices, etp_panel, "Buy & Hold 1x", pd.Series(1.0, index=idx), cfg["trading_cost"])
wrows = []
for name, lev, lev_max in winners:
    r = sw.run_one_backtest(prices, etp_panel, name, lev, cfg["trading_cost"])
    wrows.append({
        "Asset": "spx", "Strategy": name, "Leverage_Max": lev_max,
        "CAGR_pct": r["cagr"]*100, "Vol_pct": r["volatility"]*100, "Sharpe": r["sharpe"],
        "Sortino": r["sortino"], "Calmar": r["calmar"], "MaxDD_pct": r["max_drawdown"]*100,
        "End_Value": r["end_value"], "Start_Date": str(r["start_date"])[:10],
        "End_Date": str(r["end_date"])[:10], "Years": r["years"],
        "Pct_Cash_Time": r["pct_cash_time"], "Trades_Per_Year": r["trades_per_year"],
        "Total_Trades": r["total_trades"], "Avg_Leverage": r["avg_leverage"],
        "Beat_BH_Sharpe": int(r["sharpe"] > bh1["sharpe"]), "Beat_BH_Calmar": int(r["calmar"] > bh1["calmar"]),
        "Beat_BH_DD": int(r["max_drawdown"] > bh1["max_drawdown"]), "Beat_BH_CAGR": int(r["cagr"] > bh1["cagr"]),
        "Trading_Cost_Pct": cfg["trading_cost"]*100,
    })
    print(f"  {name:44s} CAGR {r['cagr']*100:6.2f} DD {r['max_drawdown']*100:7.2f} Sharpe {r['sharpe']:.3f} Calmar {r['calmar']:.3f}", flush=True)

df = pd.concat([inc, pd.DataFrame(wrows)], ignore_index=True)
df = bx.classify_all(df)
print("\nWinner classifications:")
for w in winners:
    print(f"  {w[0]:44s} -> {df[df.Strategy==w[0]].iloc[0]['Classification']}")
print("Counts:", df[df.Asset=='spx']['Classification'].value_counts().to_dict())

# ---- 4. rebuild the two SPX tabs, drop band-sweep sheets ----
wb = openpyxl.load_workbook(BASE)
for s in ["SPX Band-Sweep Water", "SPX Band-Sweep Octane", "SPX Band-Sweep Method", "SPX Band-Sweep Diagrams"]:
    if s in wb.sheetnames:
        del wb[s]
order = wb.sheetnames[:]
pos_w = order.index("S&P 500 Water"); pos_o = order.index("S&P 500 Octane")
del wb["S&P 500 Water"]; del wb["S&P 500 Octane"]
ws_w = bx.build_detail_sheet(wb, df, "spx", "Water")
ws_o = bx.build_detail_sheet(wb, df, "spx", "Octane")
wb._sheets.remove(ws_w); wb._sheets.remove(ws_o)
wb._sheets.insert(pos_w, ws_w); wb._sheets.insert(pos_o, ws_o)
wb.save(DEST)
print(f"\nWROTE {DEST}")
print("Sheets:", wb.sheetnames)
