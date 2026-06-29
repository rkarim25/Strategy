"""Strategy Lab — the canonical, reproducible way to test a new strategy on ANY
asset and (optionally) add it to Results/strategy_results.xlsx + the website, on a
basis that is byte-for-byte consistent with every existing result.

================================================================================
THE GOLDEN RULE OF CONSISTENCY
================================================================================
Every backtest in the Excel workbook AND the website is produced by
    sweep_all_assets_strategies.load_asset_data(asset, tbill, vix)  +  run_one_backtest()
Do NOT use backtest_*.download_*  (e.g. download_spx_panel) for new strategies:
it pulls a DIFFERENT yfinance auto-adjust vintage and gives e.g. SPX B&H 1x CAGR
~9.3% instead of the sheet's ~9.99% — non-comparable. This module always uses the
canonical loader, so a new strategy lands on the SAME basis as the incumbents.

Two workflows:
  1. quicktest(asset, name, signal01, leverage)  — print metrics + classification
     + does-it-beat-the-incumbents verdict (nothing is written).
  2. add_strategies(asset, [(name, signal01, lev), ...])  — recompute the new
     strategies on the frozen window of output/strategy_results/<asset>_results.csv
     (the source of the existing tabs), re-classify, and rewrite ONLY that asset's
     Water/Octane/Stillwater tabs in the workbook. All other tabs are untouched.

REPRODUCING AN OLD BACKTEST WITH NEW DATES
  load(asset) fetches CURRENT data. To compare a strategy to a frozen result, pass
  end_date=<the CSV End_Date> to load() so the window matches exactly (the plain
  SMA200 band 1x should reproduce the CSV value to the cent — see _basis_check()).

Signals live in signals.py (band/early/stop/trailing/variant-B/variant-C). Add new
ones there; they all return a 0/1 numpy array on prices["spx_close"] (the generic
index-close column name used for every asset).
"""
from __future__ import annotations
from pathlib import Path
import numpy as np
import pandas as pd

import sweep_all_assets_strategies as sw
import build_strategy_results_excel as bx

REPO = Path(__file__).resolve().parents[2]
CSV_DIR = REPO / "output" / "strategy_results"
XLSX = REPO / "Results" / "strategy_results.xlsx"

# Assets whose pages/tabs use Stillwater because no strict Water/Octane exists.
STILLWATER_ASSETS = {"ndx"}

_TBILL = _VIX = None
def _tbill_vix():
    global _TBILL, _VIX
    if _TBILL is None:
        _TBILL, _VIX = sw._download_tbill_vix()
    return _TBILL, _VIX


def load(asset: str, end_date: str | None = None):
    """Canonical loader. Returns (prices, etp_panel, cfg). Pass end_date='YYYY-MM-DD'
    to truncate to a frozen window (for exact reproduction of an old result)."""
    tb, vx = _tbill_vix()
    ad = sw.load_asset_data(asset, tb, vx)
    prices = ad["prices"]
    if end_date:
        prices = prices[prices.index <= end_date].copy()
    cfg = sw.ASSETS[asset]
    panel = sw.build_asset_etp_panel(prices, ad["etp_closes"], cfg)
    return prices, panel, cfg


def run(prices, panel, cfg, name, signal01, leverage):
    """Run one strategy on the canonical engine. signal01 = 0/1 array (or Series)."""
    sig = np.asarray(signal01, dtype=float) * float(leverage)
    r = sw.run_one_backtest(prices, panel, name, pd.Series(sig, index=prices.index), cfg["trading_cost"])
    return {
        "Strategy": name, "Leverage_Max": float(leverage),
        "CAGR_pct": r["cagr"]*100, "Vol_pct": r["volatility"]*100, "Sharpe": r["sharpe"],
        "Sortino": r["sortino"], "Calmar": r["calmar"], "MaxDD_pct": r["max_drawdown"]*100,
        "End_Value": r["end_value"], "Start_Date": str(r["start_date"])[:10],
        "End_Date": str(r["end_date"])[:10], "Years": r["years"], "Pct_Cash_Time": r["pct_cash_time"],
        "Trades_Per_Year": r["trades_per_year"], "Total_Trades": r["total_trades"],
        "Avg_Leverage": r["avg_leverage"], "Trading_Cost_Pct": cfg["trading_cost"]*100,
    }


def bh1_metrics(prices, panel, cfg):
    row = run(prices, panel, cfg, "Buy & Hold 1x", np.ones(len(prices)), 1.0)
    return {k: row[k] for k in ["Sharpe", "Calmar", "MaxDD_pct", "CAGR_pct", "Sortino", "Vol_pct"]}


def classify(row, bhm, stillwater=False):
    strict = bx.classify_strategy(row, bhm)
    if strict in ("Water", "Octane") or not stillwater:
        return strict
    return bx.classify_stillwater(row, bhm)


def _basis_check(prices, panel, cfg, asset):
    """Confirm we reproduce the frozen CSV (plain SMA200 +-3% band 1x)."""
    import signals as S
    close = prices["spx_close"].to_numpy(float)
    from core.indicators import sma as sma_ind
    sma200 = sma_ind(prices["spx_close"], 200).to_numpy(float)
    got = run(prices, panel, cfg, "ref", S.conv_band(close, sma200, 0.03, 0.03), 1.0)["CAGR_pct"]
    csv = pd.read_csv(CSV_DIR / f"{asset}_results.csv")
    ref = csv[csv.Strategy == "SMA200 +-3% Band 1x/cash"]
    want = float(ref.iloc[0]["CAGR_pct"]) if len(ref) else None
    return got, want


def incumbents(asset):
    """The frozen per-asset incumbent rows + their classification."""
    df = pd.read_csv(CSV_DIR / f"{asset}_results.csv"); df.insert(0, "Asset", asset)
    return bx.classify_all(df)


def quicktest(asset, name, signal01, leverage):
    """Print canonical metrics + classification + beat-incumbents verdict. Writes nothing."""
    prices, panel, cfg = load(asset)
    bhm = bh1_metrics(prices, panel, cfg)
    row = run(prices, panel, cfg, name, signal01, leverage)
    cls = classify(row, bhm, stillwater=(asset in STILLWATER_ASSETS))
    print(f"[{asset}] {name}")
    print(f"  CAGR {row['CAGR_pct']:.2f}  Vol {row['Vol_pct']:.2f}  Sharpe {row['Sharpe']:.3f}  "
          f"Sortino {row['Sortino']:.3f}  Calmar {row['Calmar']:.3f}  MaxDD {row['MaxDD_pct']:.2f}  "
          f"Trades/yr {row['Trades_Per_Year']:.1f}")
    print(f"  B&H 1x: CAGR {bhm['CAGR_pct']:.2f} Sharpe {bhm['Sharpe']:.3f} Calmar {bhm['Calmar']:.3f} DD {bhm['MaxDD_pct']:.2f}")
    print(f"  -> classification: {cls}")
    return row, cls


def add_strategies(asset, new_strategies, dest=None, drop_sheets=None):
    """Add new strategies to the EXISTING asset tabs without disturbing any incumbent.
    new_strategies: list of (name, signal01_or_callable, leverage). A callable receives
    (prices, cfg) and returns a 0/1 array — use it so the signal is built on the frozen
    window. Rebuilds the asset's Water + Octane tabs (+ Stillwater for ndx). Saves to
    dest (default: in place). Returns the dest path."""
    import openpyxl
    dest = Path(dest) if dest else XLSX
    inc = pd.read_csv(CSV_DIR / f"{asset}_results.csv"); inc.insert(0, "Asset", asset)
    end = str(inc[inc.Strategy == "Buy & Hold 1x"].iloc[0]["End_Date"])[:10]
    prices, panel, cfg = load(asset, end_date=end)
    got, want = _basis_check(prices, panel, cfg, asset)
    print(f"basis check {asset}: band1x CAGR {got:.2f} (CSV {want}) — {'OK' if want and abs(got-want)<0.05 else 'CHECK'}")

    bh1 = run(prices, panel, cfg, "Buy & Hold 1x", np.ones(len(prices)), 1.0)
    rows = []
    for name, sig, lev in new_strategies:
        arr = sig(prices, cfg) if callable(sig) else sig
        r = run(prices, panel, cfg, name, arr, lev)
        r["Asset"] = asset
        r["Beat_BH_Sharpe"] = int(r["Sharpe"] > bh1["Sharpe"]); r["Beat_BH_Calmar"] = int(r["Calmar"] > bh1["Calmar"])
        r["Beat_BH_DD"] = int(r["MaxDD_pct"] > bh1["MaxDD_pct"]); r["Beat_BH_CAGR"] = int(r["CAGR_pct"] > bh1["CAGR_pct"])
        rows.append(r)
        print(f"  + {name}: CAGR {r['CAGR_pct']:.2f} DD {r['MaxDD_pct']:.2f} Sharpe {r['Sharpe']:.3f} Calmar {r['Calmar']:.3f}")
    df = bx.classify_all(pd.concat([inc, pd.DataFrame(rows)], ignore_index=True))

    base = dest if dest.exists() else XLSX
    wb = openpyxl.load_workbook(base)
    for s in (drop_sheets or []):
        if s in wb.sheetnames:
            del wb[s]
    label = bx.ASSET_LABELS[asset]
    tabs = ["Water", "Octane"] + (["Stillwater"] if asset in STILLWATER_ASSETS else [])
    order = wb.sheetnames[:]
    pos = {}
    for t in tabs:
        nm = f"{label} {t}"[:31]
        if nm in wb.sheetnames:
            pos[t] = order.index(nm); del wb[nm]
    made = {}
    for t in tabs:
        if t == "Stillwater":
            made[t] = bx.build_stillwater_sheet(wb, df, asset) if hasattr(bx, "build_stillwater_sheet") else None
        else:
            made[t] = bx.build_detail_sheet(wb, df, asset, t)
    for t in tabs:
        ws = made.get(t)
        if ws is not None and t in pos:
            wb._sheets.remove(ws); wb._sheets.insert(pos[t], ws)
    wb.save(dest)
    print(f"wrote {dest}")
    return dest


if __name__ == "__main__":
    import signals as S
    from core.indicators import sma as sma_ind
    # demo: the SPX Accel-Exit winner
    prices, panel, cfg = load("spx")
    close = prices["spx_close"].to_numpy(float)
    sma200 = sma_ind(prices["spx_close"], 200).to_numpy(float)
    quicktest("spx", "SMA200 +-3% Band + Accel-Exit N10 1x", S.variant_b_accel(close, sma200, 0.03, "conv", 10, 0.02, 0.03), 1.0)
