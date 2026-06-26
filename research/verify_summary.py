import sys as _s, pathlib as _p; _s.path.insert(0, str(_p.Path(__file__).resolve().parent.parent))  # repo root importable (moved into research/)
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Results/strategy_results.xlsx')

# Check Summary sheet for SPX and NDX counts in detail
ws = wb['Summary']
print('=== SUMMARY SHEET (all rows, all cols) ===')
for row in range(1, ws.max_row+1):
    vals = []
    for col in range(1, ws.max_column+1):
        v = ws.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
        else:
            vals.append("")
    if any(vals):  # Only print rows with at least one non-empty cell
        print(f'  Row {row}: {" | ".join(vals)}')
print()

# Check the Definitions sheet to understand the classification
ws_defs = wb['Definitions']
print('=== DEFINITIONS SHEET ===')
for row in range(1, min(20, ws_defs.max_row+1)):
    vals = []
    for col in range(1, min(10, ws_defs.max_column+1)):
        v = ws_defs.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
        else:
            vals.append("")
    if any(vals):  # Only print rows with at least one non-empty cell
        print(f'  Row {row}: {" | ".join(vals)}')
print()