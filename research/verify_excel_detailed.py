import sys as _s, pathlib as _p; _s.path.insert(0, str(_p.Path(__file__).resolve().parent.parent))  # repo root importable (moved into research/)
import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('Results/strategy_results.xlsx')

# Check Nasdaq 100 Water sheet - more detailed view
ws4 = wb['Nasdaq 100 Water']
print(f'=== Nasdaq 100 Water sheet: {ws4.max_row} rows, {ws4.max_column} cols ===')
# Print more rows to see if there's data beyond row 10
for row in range(1, min(20, ws4.max_row+1)):
    vals = []
    for col in range(1, min(10, ws4.max_column+1)):
        v = ws4.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals:
        print(f'  Row {row}: {" | ".join(vals)}')
print()

# Check Nasdaq 100 Octane sheet - more detailed view
ws5 = wb['Nasdaq 100 Octane']
print(f'=== Nasdaq 100 Octane sheet: {ws5.max_row} rows, {ws5.max_column} cols ===')
# Print more rows to see if there's data beyond row 10
for row in range(1, min(20, ws5.max_row+1)):
    vals = []
    for col in range(1, min(10, ws5.max_column+1)):
        v = ws5.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals:
        print(f'  Row {row}: {" | ".join(vals)}')
print()

# Let's also check if there are any strategies with non-zero values in Nasdaq 100 sheets
print("=== Checking for non-zero strategy data in Nasdaq 100 sheets ===")

# Check for any non-header, non-empty rows in Nasdaq 100 Water
print("Nasdaq 100 Water - Non-empty rows beyond header:")
for row in range(11, min(50, ws4.max_row+1)):  # Start after row 10
    vals = []
    for col in range(1, min(10, ws4.max_column+1)):
        v = ws4.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals and any(v.strip() != "" for v in vals):
        print(f'  Row {row}: {" | ".join(vals)}')

print()

# Check for any non-header, non-empty rows in Nasdaq 100 Octane
print("Nasdaq 100 Octane - Non-empty rows beyond header:")
for row in range(11, min(50, ws5.max_row+1)):  # Start after row 10
    vals = []
    for col in range(1, min(10, ws5.max_column+1)):
        v = ws5.cell(row=row, column=col).value
        if v is not None:
            vals.append(str(v)[:50])
    if vals and any(v.strip() != "" for v in vals):
        print(f'  Row {row}: {" | ".join(vals)}')