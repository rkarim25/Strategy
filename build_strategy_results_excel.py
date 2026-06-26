"""Build comprehensive Strategy Results Excel workbook with Water/Octane classification.

Reads 11 CSV files from output/strategy_results/ and produces a formatted
Excel workbook at Results/strategy_results.xlsx with:
  - 1 Summary sheet
  - 20 detail sheets (10 assets x 2 classifications: Water + Octane)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from pathlib import Path
import os
import traceback
import sys

# ── Paths ──────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
CSV_DIR = ROOT / 'output' / 'strategy_results'
RESULTS_DIR = ROOT / 'Results'
EXCEL_PATH = RESULTS_DIR / 'strategy_results.xlsx'

# ── Asset label mapping ───────────────────────────────────────────────────
ASSET_LABELS = {
    'spx':        'S&P 500',
    'spxew':      'S&P 500 EW',
    'ndx':        'Nasdaq 100',
    'rut':        'Russell 2000',
    'gold':       'Gold',
    'tlt':        '20Y+ Treasuries',
    'ftse250':    'FTSE 250',
    'dax':        'DAX',
    'msci_em':    'MSCI EM',
    'msci_world': 'MSCI World',
}

# Order for display
ASSET_ORDER = ['spx', 'spxew', 'ndx', 'rut', 'gold', 'tlt',
               'ftse250', 'dax', 'msci_em', 'msci_world']

# Trading cost labels for display
TRADING_COST_LABELS = {
    'spx': '0.10%', 'spxew': '0.12%', 'ndx': '0.10%', 'rut': '0.15%',
    'gold': '0.15%', 'tlt': '0.10%', 'ftse250': '0.20%', 'dax': '0.20%',
    'msci_em': '0.20%', 'msci_world': '0.15%'
}

# ── Column sets ───────────────────────────────────────────────────────────
# Full detail columns (Water/Octane expanded sections)
DETAIL_COLS = [
    'Strategy', 'Leverage_Max', 'CAGR_pct', 'Vol_pct', 'Sharpe', 'Sortino',
    'Calmar', 'MaxDD_pct', 'End_Value', 'Start_Date', 'End_Date', 'Years',
    'Pct_Cash_Time', 'Trades_Per_Year', 'Total_Trades', 'Avg_Leverage',
]

# Collapsed "Neither" columns
NEITHER_COLS = ['Strategy', 'MaxDD_pct']

# Summary columns
SUMMARY_COLS = [
    'Asset', 'Total_Strategies', 'Water_Count', 'Octane_Count',
    'Neither_Count', 'Best_Water_Strategy', 'Best_Water_CAGR',
    'Best_Octane_Strategy', 'Best_Octane_Calmar',
    'BH_1x_CAGR', 'BH_1x_MaxDD',
]

# ── Number format mapping ─────────────────────────────────────────────────
# (column_name, openpyxl number_format)
NUMBER_FORMATS = {
    'CAGR_pct':        '0.00',
    'Vol_pct':         '0.00',
    'MaxDD_pct':       '0.00',
    'Pct_Cash_Time':   '0.00',
    'Sharpe':          '0.000',
    'Sortino':         '0.000',
    'Calmar':          '0.000',
    'Avg_Leverage':    '0.000',
    'Trades_Per_Year': '0.000',
    'End_Value':       '$#,##0',
    'Total_Trades':    '0',
    'Years':           '0.0',
    'Leverage_Max':    '0.0',
    'BH_1x_CAGR':      '0.00',
    'BH_1x_MaxDD':     '0.00',
    'Best_Water_CAGR': '0.00',
    'Best_Octane_Calmar': '0.000',
}

# ── Styles ─────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
BH_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # light blue
BH_FONT = Font(name='Calibri', size=11, bold=True)
WATER_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # light green
OCTANE_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # light green
NEITHER_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')  # light gray
NEITHER_FONT = Font(name='Calibri', size=11, italic=True, color='808080')
NORMAL_FONT = Font(name='Calibri', size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)


def load_all_data():
    """Load the combined CSV and return a DataFrame with Asset column."""
    csv_path = CSV_DIR / 'all_assets_combined.csv'
    if not csv_path.exists():
        # Fall back to reading individual CSVs and combining
        print("Combined CSV not found, building from individual CSVs...", flush=True)
        frames = []
        for asset_key in ASSET_ORDER:
            fpath = CSV_DIR / f'{asset_key}_results.csv'
            if fpath.exists():
                df = pd.read_csv(fpath)
                df.insert(0, 'Asset', asset_key)
                frames.append(df)
                print(f"  Loaded {asset_key}: {len(df)} rows", flush=True)
        if not frames:
            raise FileNotFoundError(f"No CSV files found in {CSV_DIR}")
        df = pd.concat(frames, ignore_index=True)
    else:
        df = pd.read_csv(csv_path)
        print(f"Loaded combined CSV: {len(df)} rows, {df['Asset'].nunique()} assets", flush=True)

    # Ensure numeric columns are numeric
    numeric_cols = ['Leverage_Max', 'CAGR_pct', 'Vol_pct', 'Sharpe', 'Sortino',
                    'Calmar', 'MaxDD_pct', 'End_Value', 'Years', 'Pct_Cash_Time',
                    'Trades_Per_Year', 'Total_Trades', 'Avg_Leverage',
                    'Beat_BH_Sharpe', 'Beat_BH_Calmar', 'Beat_BH_DD', 'Beat_BH_CAGR']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    return df


def is_bh_row(strategy_name):
    """Check if a row is a Buy & Hold benchmark."""
    return isinstance(strategy_name, str) and strategy_name.startswith('Buy & Hold')


def classify_strategy(row, bh_metrics=None):
    """Classify a single strategy row as 'Water', 'Octane', or 'Neither'.

    Water: A strategy qualifies as Water if:
    1. It does NOT sacrifice ANY metric relative to B&H 1x — meaning ALL of these must be true:
       - Sharpe >= bh_sharpe (Sharpe not worse)
       - Calmar >= bh_calmar (Calmar not worse)
       - MaxDD_pct >= bh_maxdd (MaxDD not worse, i.e., drawdown is same or smaller)
       - CAGR_pct >= bh_cagr (CAGR not worse)
       - Sortino >= bh_sortino (Sortino not worse)
       - Vol_pct <= bh_vol (Vol not worse, lower is better)
    2. AND at least one of these is strictly better:
       - CAGR_pct > bh_cagr (CAGR strictly better)
       - MaxDD_pct > bh_maxdd (MaxDD strictly better)

    Octane: A strategy qualifies as Octane if:
    1. CAGR_pct > bh_cagr (CAGR strictly better than B&H 1x)
    2. AND Calmar > bh_calmar (Calmar strictly better than B&H 1x)
    3. AND MaxDD_pct >= -45.0 (max drawdown no worse than -45%)
    4. AND Trades_Per_Year <= 30.0 (max 30 trades per year)

    If both qualify → Water (more prestigious).
    """
    if is_bh_row(row.get('Strategy', '')):
        return 'Benchmark'

    # If no B&H metrics provided, fall back to old method
    if bh_metrics is None:
        water = (int(row.get('Beat_BH_Sharpe', 0)) == 1
                 and int(row.get('Beat_BH_Calmar', 0)) == 1
                 and int(row.get('Beat_BH_DD', 0)) == 1
                 and int(row.get('Beat_BH_CAGR', 0)) == 1)

        octane = (int(row.get('Beat_BH_Calmar', 0)) == 1
                  and float(row.get('MaxDD_pct', -999)) >= -45.0
                  and float(row.get('Trades_Per_Year', 999)) <= 30.0)

        if water:
            return 'Water'
        elif octane:
            return 'Octane'
        else:
            return 'Neither'

    # Use direct metric comparison
    bh_sharpe = bh_metrics['Sharpe']
    bh_calmar = bh_metrics['Calmar']
    bh_maxdd = bh_metrics['MaxDD_pct']
    bh_cagr = bh_metrics['CAGR_pct']
    bh_sortino = bh_metrics['Sortino']
    bh_vol = bh_metrics['Vol_pct']

    # Get strategy metrics
    strat_sharpe = float(row.get('Sharpe', 0))
    strat_calmar = float(row.get('Calmar', 0))
    strat_maxdd = float(row.get('MaxDD_pct', 0))
    strat_cagr = float(row.get('CAGR_pct', 0))
    strat_sortino = float(row.get('Sortino', 0))
    strat_vol = float(row.get('Vol_pct', 0))
    strat_trades = float(row.get('Trades_Per_Year', 999))

    # Water: no metric worse than B&H 1x, AND at least one of CAGR/MaxDD strictly better
    not_worse = (
        strat_sharpe >= bh_sharpe and
        strat_calmar >= bh_calmar and
        strat_maxdd >= bh_maxdd and  # MaxDD is negative, so >= means smaller drawdown
        strat_cagr >= bh_cagr and
        strat_sortino >= bh_sortino and
        strat_vol <= bh_vol  # lower vol is better
    )
    strictly_better = (
        strat_cagr > bh_cagr or
        strat_maxdd > bh_maxdd  # smaller drawdown is better
    )
    is_water = not_worse and strictly_better

    # Octane: CAGR > B&H 1x AND Calmar > B&H 1x AND MaxDD >= -45% AND Trades <= 30
    is_octane = (
        strat_cagr > bh_cagr and
        strat_calmar > bh_calmar and
        strat_maxdd >= -45.0 and
        strat_trades <= 30.0
    )

    if is_water:
        return 'Water'
    elif is_octane:
        return 'Octane'
    else:
        return 'Neither'


# Percentage points of CAGR a Stillwater strategy may give up versus B&H 1x.
STILLWATER_CAGR_TOL = 2.0


def classify_stillwater(row, bh_metrics, tol=STILLWATER_CAGR_TOL):
    """Stillwater: the Water/Octane risk criteria, but CAGR may be up to `tol`
    percentage points BELOW Buy & Hold 1x (instead of >= / > B&H). Intended for
    assets such as Nasdaq 100 where strict Water/Octane is unreachable because the
    benchmark's high CAGR is only achieved through an extreme (-82%) drawdown.

    Returns 'Stillwater-Water', 'Stillwater-Octane', 'Neither', or 'Benchmark'.
      Water-style : no risk metric worse than B&H 1x (Sharpe, Sortino, Calmar,
                    MaxDD, Vol), CAGR >= B&H - tol, and >=1 metric strictly better.
      Octane-style: CAGR >= B&H - tol, Calmar > B&H, MaxDD >= -45%, Trades/yr <= 30.
    Water-style takes precedence when both qualify.
    """
    if is_bh_row(row.get('Strategy', '')):
        return 'Benchmark'
    if bh_metrics is None:
        return 'Neither'

    cagr = float(row.get('CAGR_pct', 0))
    dd = float(row.get('MaxDD_pct', 0))
    cal = float(row.get('Calmar', 0))
    sh = float(row.get('Sharpe', 0))
    so = float(row.get('Sortino', 0))
    vol = float(row.get('Vol_pct', 0))
    trades = float(row.get('Trades_Per_Year', 999))

    bh_cagr = bh_metrics['CAGR_pct']
    cagr_ok = cagr >= bh_cagr - tol  # relaxed CAGR floor

    sw_water = (
        sh >= bh_metrics['Sharpe'] and so >= bh_metrics['Sortino'] and
        cal >= bh_metrics['Calmar'] and dd >= bh_metrics['MaxDD_pct'] and
        vol <= bh_metrics['Vol_pct'] and cagr_ok and
        (dd > bh_metrics['MaxDD_pct'] or cagr > bh_cagr or
         sh > bh_metrics['Sharpe'] or cal > bh_metrics['Calmar'])
    )
    sw_octane = (cagr_ok and cal > bh_metrics['Calmar'] and dd >= -45.0 and trades <= 30.0)

    if sw_water:
        return 'Stillwater-Water'
    if sw_octane:
        return 'Stillwater-Octane'
    return 'Neither'


def classify_all(df):
    """Add classification column to dataframe."""
    df = df.copy()
    
    # Extract B&H 1x metrics for each asset and classify strategies
    classifications = []
    for idx, row in df.iterrows():
        asset_key = row['Asset']
        
        # Find B&H 1x row for this asset
        asset_df = df[df['Asset'] == asset_key]
        bh1_row = asset_df[asset_df['Strategy'] == 'Buy & Hold 1x']
        if len(bh1_row) > 0:
            bh_metrics = {
                'Sharpe': float(bh1_row.iloc[0]['Sharpe']),
                'Calmar': float(bh1_row.iloc[0]['Calmar']),
                'MaxDD_pct': float(bh1_row.iloc[0]['MaxDD_pct']),
                'CAGR_pct': float(bh1_row.iloc[0]['CAGR_pct']),
                'Sortino': float(bh1_row.iloc[0]['Sortino']),
                'Vol_pct': float(bh1_row.iloc[0]['Vol_pct'])
            }
        else:
            bh_metrics = None
        
        # Classify this strategy
        classification = classify_strategy(row, bh_metrics)
        classifications.append(classification)
    
    df['Classification'] = classifications
    return df


def build_summary_sheet(wb, df):
    """Build the Summary sheet with per-asset overview."""
    ws = wb.create_sheet(title='Summary')

    # Build summary data
    summary_rows = []
    for asset_key in ASSET_ORDER:
        asset_label = ASSET_LABELS[asset_key]
        asset_df = df[df['Asset'] == asset_key]

        # Exclude B&H rows from strategy counts
        strategies_df = asset_df[~asset_df['Strategy'].apply(is_bh_row)]
        total = len(strategies_df)
        water_count = (strategies_df['Classification'] == 'Water').sum()
        octane_count = (strategies_df['Classification'] == 'Octane').sum()
        neither_count = (strategies_df['Classification'] == 'Neither').sum()

        # Best Water strategy (by CAGR)
        water_df = strategies_df[strategies_df['Classification'] == 'Water']
        if len(water_df) > 0:
            best_water = water_df.loc[water_df['CAGR_pct'].idxmax()]
            best_water_name = best_water['Strategy']
            best_water_cagr = best_water['CAGR_pct']
        else:
            best_water_name = '—'
            best_water_cagr = None

        # Best Octane strategy (by Calmar, excluding Water)
        octane_df = strategies_df[strategies_df['Classification'] == 'Octane']
        if len(octane_df) > 0:
            best_octane = octane_df.loc[octane_df['Calmar'].idxmax()]
            best_octane_name = best_octane['Strategy']
            best_octane_calmar = best_octane['Calmar']
        else:
            best_octane_name = '—'
            best_octane_calmar = None

        # B&H 1x stats
        bh_1x = asset_df[asset_df['Strategy'] == 'Buy & Hold 1x']
        if len(bh_1x) > 0:
            bh_1x_cagr = bh_1x.iloc[0]['CAGR_pct']
            bh_1x_maxdd = bh_1x.iloc[0]['MaxDD_pct']
        else:
            bh_1x_cagr = None
            bh_1x_maxdd = None

        summary_rows.append({
            'Asset': asset_label,
            'Total_Strategies': total,
            'Water_Count': water_count,
            'Octane_Count': octane_count,
            'Neither_Count': neither_count,
            'Best_Water_Strategy': best_water_name,
            'Best_Water_CAGR': best_water_cagr,
            'Best_Octane_Strategy': best_octane_name,
            'Best_Octane_Calmar': best_octane_calmar,
            'BH_1x_CAGR': bh_1x_cagr,
            'BH_1x_MaxDD': bh_1x_maxdd,
        })

    # Write header
    for col_idx, col_name in enumerate(SUMMARY_COLS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Write data
    for row_idx, row_data in enumerate(summary_rows, 2):
        for col_idx, col_name in enumerate(SUMMARY_COLS, 1):
            val = row_data.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val if val is not None else '')
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center' if col_idx > 1 else 'left')

            # Number formatting
            fmt = NUMBER_FORMATS.get(col_name)
            if fmt and val is not None:
                cell.number_format = fmt

    # Freeze panes
    ws.freeze_panes = 'A2'

    # Auto-fit column widths
    auto_fit_columns(ws)

    # Conditional formatting: highlight best values
    # Best Water CAGR (column 7)
    if len(summary_rows) > 1:
        last_data_row = len(summary_rows) + 1
        ws.conditional_formatting.add(
            f'G2:G{last_data_row}',
            FormulaRule(
                formula=[f'G2=MAX($G$2:$G${last_data_row})'],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                font=Font(bold=True)
            )
        )
        # Best Octane Calmar (column 9)
        ws.conditional_formatting.add(
            f'I2:I{last_data_row}',
            FormulaRule(
                formula=[f'I2=MAX($I$2:$I${last_data_row})'],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                font=Font(bold=True)
            )
        )

    # Add trading cost assumptions
    current_row = len(summary_rows) + 3  # Skip header + data rows + 1 blank
    
    # Trading cost assumptions header
    cell = ws.cell(row=current_row, column=1)
    cell.value = "Trading Cost Assumptions (from mid-price):"
    cell.font = Font(name='Calibri', size=11, bold=True, italic=True)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(SUMMARY_COLS))
    current_row += 1
    
    # Trading cost assumptions details
    cost_text = ("S&P 500: 0.10% | S&P 500 EW: 0.12% | Nasdaq 100: 0.10% | Russell 2000: 0.15% | "
                 "Gold: 0.15% | 20Y+ Treasuries: 0.10% | FTSE 250: 0.20% | DAX: 0.20% | "
                 "MSCI EM: 0.20% | MSCI World: 0.15%")
    cell = ws.cell(row=current_row, column=1)
    cell.value = cost_text
    cell.font = Font(name='Calibri', size=9, italic=True, color='808080')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(SUMMARY_COLS))

    # Stillwater note (Nasdaq 100): count strategies passing the relaxed-CAGR class
    ndx_df = df[df['Asset'] == 'ndx']
    ndx_bh = ndx_df[ndx_df['Strategy'] == 'Buy & Hold 1x']
    sw_count = 0
    if len(ndx_bh) > 0:
        ndx_bh_metrics = {k: float(ndx_bh.iloc[0][k]) for k in
                          ['Sharpe', 'Calmar', 'MaxDD_pct', 'CAGR_pct', 'Sortino', 'Vol_pct']}
        ndx_strats = ndx_df[~ndx_df['Strategy'].apply(is_bh_row)]
        sw_count = int(ndx_strats.apply(
            lambda r: classify_stillwater(r, ndx_bh_metrics) in ('Stillwater-Water', 'Stillwater-Octane'),
            axis=1).sum())
    current_row += 1
    cell = ws.cell(row=current_row, column=1)
    cell.value = (f"Nasdaq 100 Stillwater (Water/Octane risk profile, CAGR within "
                  f"{STILLWATER_CAGR_TOL:.0f}pp of B&H 1x): {sw_count} strategies — see 'Nasdaq 100 Stillwater' sheet.")
    cell.font = Font(name='Calibri', size=9, italic=True, color='808080')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(SUMMARY_COLS))

    print(f"  Summary sheet: {len(summary_rows)} assets", flush=True)
    return ws


def build_definitions_sheet(wb):
    """Build the Definitions sheet with Water/Octane classification definitions and methodology."""
    ws = wb.create_sheet(title="Definitions", index=0)  # Make it the first sheet
    
    # Define styles
    title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    header_font = Font(name='Calibri', size=12, bold=True, color='4472C4')
    normal_font = Font(name='Calibri', size=11, bold=False, color='000000')
    
    # Title row
    title_cell = ws.cell(row=1, column=1, value="Strategy Classification Definitions & Methodology")
    title_cell.font = title_font
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    
    # Running row counter so sections can be added without manual renumbering.
    r = 3

    def line(text, font=normal_font):
        nonlocal r
        ws.cell(row=r, column=1, value=text).font = font
        r += 1

    def gap():
        nonlocal r
        r += 1

    # Water Classification section
    line("Water Classification", header_font)
    line("A strategy qualifies as Water if it improves on Buy & Hold 1x on at least one of CAGR or MaxDD, without sacrificing any metric relative to B&H 1x.")
    line("Specifically:")
    line("  - CAGR_pct >= B&H 1x CAGR_pct (not worse)")
    line("  - MaxDD_pct >= B&H 1x MaxDD_pct (not worse, i.e., smaller or equal drawdown)")
    line("  - Sharpe >= B&H 1x Sharpe (not worse)")
    line("  - Calmar >= B&H 1x Calmar (not worse)")
    line("  - Sortino >= B&H 1x Sortino (not worse)")
    line("  - Vol_pct <= B&H 1x Vol_pct (not worse, i.e., lower or equal volatility)")
    line("  AND at least one of the following is strictly better:")
    line("  - CAGR_pct > B&H 1x CAGR_pct (strictly better CAGR)")
    line("  - MaxDD_pct > B&H 1x MaxDD_pct (strictly better drawdown, i.e., smaller)")
    line("Water takes precedence over Octane if a strategy qualifies for both.")
    gap()

    # Octane Classification section
    line("Octane Classification", header_font)
    line("A strategy qualifies as Octane if ALL of the following are true:")
    line("  - CAGR_pct > B&H 1x CAGR_pct (strictly better CAGR)")
    line("  - Calmar > B&H 1x Calmar (strictly better Calmar ratio)")
    line("  - MaxDD_pct >= -45.0% (maximum drawdown no worse than -45%)")
    line("  - Trades_Per_Year <= 30.0 (maximum 30 trades per year)")
    gap()

    # Stillwater Classification section (Nasdaq 100)
    line("Stillwater Classification (Nasdaq 100)", header_font)
    line("A relaxed class for assets where strict Water/Octane is unreachable because the benchmark's high CAGR")
    line("is only achieved through an extreme drawdown (Nasdaq 100 B&H 1x: ~16.5% CAGR at -82.5% MaxDD).")
    line(f"Stillwater keeps the Water/Octane RISK criteria but lets CAGR fall up to {STILLWATER_CAGR_TOL:.0f} percentage points below B&H 1x:")
    line(f"  - Water-style : no risk metric worse than B&H 1x (Sharpe, Sortino, Calmar, MaxDD, Vol),")
    line(f"                  CAGR >= B&H 1x CAGR - {STILLWATER_CAGR_TOL:.0f}pp, and at least one metric strictly better.")
    line(f"  - Octane-style: CAGR >= B&H 1x CAGR - {STILLWATER_CAGR_TOL:.0f}pp, Calmar > B&H 1x, MaxDD >= -45%, Trades/yr <= 30.")
    line("Shown only on the 'Nasdaq 100 Stillwater' sheet. Captures defensive variants that give up <=2pp of CAGR")
    line("to roughly halve the drawdown (e.g. SMA50/200 Golden Cross 1x: ~15.9% CAGR at -40% MaxDD, Calmar ~0.40).")
    gap()

    # Trading Cost Assumptions section
    line("Trading Cost Assumptions (from mid-price)", header_font)
    line("S&P 500: 0.10% | S&P 500 EW: 0.12% | Nasdaq 100: 0.10% | Russell 2000: 0.15%")
    line("Gold: 0.15% | 20Y+ Treasuries: 0.10% | FTSE 250: 0.20% | DAX: 0.20%")
    line("MSCI EM: 0.20% | MSCI World: 0.15%")
    line("Costs are calibrated to real ETF bid-ask half-spreads with a conservative buffer (2-10x actual spread).")
    gap()

    # Backtest Methodology section
    line("Backtest Methodology", header_font)
    line("Engine: PortfolioEngine (engine.py) with 1-day signal delay (no look-ahead bias)")
    line("Funding: VIX-linked borrow spread (0.6% base + 30bp/10pts above VIX 15, cap 2.6%, +20bp at 3x)")
    line("ETP Returns: Real listed ETP daily returns (SPY/SSO/UPRO, QQQ/QLD/TQQQ, etc.) with synthetic daily-reset fallback")
    line("Annual Inflow: $10 absolute on $100 initial capital")
    line("No drawdown protection (max_drawdown_limit=None, hard_drawdown_floor=False)")
    line("Time Period: Full available history per asset (S&P 500 from 1950, others from inception)")
    line("Benchmarks: Buy & Hold 1x, 2x, 3x (where leverage ETPs exist)")
    gap()

    # Data Sources section
    line("Data Sources", header_font)
    line("Yahoo Finance via yfinance: ^GSPC, ^IRX, ^VIX, ^NDX, ^RUT, ^GDAXI, ^FTMC")
    line("ETFs: SPY/SSO/UPRO, QQQ/QLD/TQQQ, IWM/UWM/TNA, GLD/UGL, TLT/UBT/TMF, RSP, EEM, SWDA.L")

    # Set column width for readability
    ws.column_dimensions['A'].width = 120
    
    print("  Definitions sheet created", flush=True)
    return ws


def build_detail_sheet(wb, df, asset_key, classification):
    """Build a Water or Octane detail sheet for a single asset.

    Args:
        classification: 'Water' or 'Octane'
    """
    asset_label = ASSET_LABELS[asset_key]
    sheet_name = f'{asset_label} {classification}'
    # Truncate sheet name to 31 chars (Excel limit)
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    ws = wb.create_sheet(title=sheet_name)

    asset_df = df[df['Asset'] == asset_key].copy()

    # Separate B&H rows, qualifying rows, and neither rows
    bh_rows = asset_df[asset_df['Strategy'].apply(is_bh_row)]
    strategies_df = asset_df[~asset_df['Strategy'].apply(is_bh_row)]

    if classification == 'Water':
        qualifying = strategies_df[strategies_df['Classification'] == 'Water']
        # Neither = everything not Water (including Octane and Neither)
        neither = strategies_df[strategies_df['Classification'] != 'Water']
        sort_col = 'Sharpe'
        sort_ascending = False
    else:  # Octane
        # Exclude strategies already classified as Water
        qualifying = strategies_df[strategies_df['Classification'] == 'Octane']
        neither = strategies_df[~strategies_df['Classification'].isin(['Water', 'Octane'])]
        sort_col = 'Calmar'
        sort_ascending = False

    # Sort qualifying strategies
    qualifying = qualifying.sort_values(sort_col, ascending=sort_ascending)
    # Sort neither by MaxDD (worst first = most negative first)
    neither = neither.sort_values('MaxDD_pct', ascending=True)

    current_row = 1

    # ── Section A header ──
    section_label = f'{asset_label} — {classification} Strategies'
    cell = ws.cell(row=current_row, column=1, value=section_label)
    cell.font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(DETAIL_COLS))
    current_row += 1
    
    # Add trading cost assumption right after section header
    cost_label = TRADING_COST_LABELS.get(asset_key, '0.50%')  # Default to 0.50% if not found
    cost_text = f"Trading cost assumption: {cost_label} from mid-price"
    cell = ws.cell(row=current_row, column=1, value=cost_text)
    cell.font = Font(name='Calibri', size=9, italic=True, color='808080')
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(DETAIL_COLS))
    current_row += 1

    # ── B&H benchmark rows ──
    cell = ws.cell(row=current_row, column=1, value='Benchmarks (Buy & Hold):')
    cell.font = Font(name='Calibri', size=11, bold=True, italic=True)
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(DETAIL_COLS))
    current_row += 1

    # Write detail header
    for col_idx, col_name in enumerate(DETAIL_COLS, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    current_row += 1

    # Write B&H rows
    bh_sorted = bh_rows.sort_values('Leverage_Max', ascending=True)
    for _, bh_row in bh_sorted.iterrows():
        for col_idx, col_name in enumerate(DETAIL_COLS, 1):
            val = bh_row.get(col_name)
            cell = ws.cell(row=current_row, column=col_idx, value=val if pd.notna(val) else '')
            cell.font = BH_FONT
            cell.fill = BH_FILL
            cell.border = THIN_BORDER
            fmt = NUMBER_FORMATS.get(col_name)
            if fmt and pd.notna(val):
                cell.number_format = fmt
        current_row += 1

    # ── Qualifying strategies ──
    current_row += 1  # blank separator row
    qual_label = f'{classification} Strategies ({len(qualifying)}):'
    cell = ws.cell(row=current_row, column=1, value=qual_label)
    cell.font = Font(name='Calibri', size=11, bold=True, color='375623')
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(DETAIL_COLS))
    current_row += 1

    # Re-write header for qualifying section
    for col_idx, col_name in enumerate(DETAIL_COLS, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    current_row += 1

    qual_start_row = current_row
    for _, q_row in qualifying.iterrows():
        for col_idx, col_name in enumerate(DETAIL_COLS, 1):
            val = q_row.get(col_name)
            cell = ws.cell(row=current_row, column=col_idx, value=val if pd.notna(val) else '')
            cell.font = NORMAL_FONT
            cell.fill = WATER_FILL if classification == 'Water' else OCTANE_FILL
            cell.border = THIN_BORDER
            fmt = NUMBER_FORMATS.get(col_name)
            if fmt and pd.notna(val):
                cell.number_format = fmt
        current_row += 1
    qual_end_row = current_row - 1

    # ── Neither strategies (collapsed) ──
    current_row += 1  # blank separator
    neither_label = f'Neither Strategies ({len(neither)}) — collapsed (Strategy + MaxDD only):'
    cell = ws.cell(row=current_row, column=1, value=neither_label)
    cell.font = Font(name='Calibri', size=11, bold=True, color='808080')
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=len(NEITHER_COLS))
    current_row += 1

    # Neither header
    for col_idx, col_name in enumerate(NEITHER_COLS, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = Font(name='Calibri', size=11, bold=True, color='595959')
        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        cell.border = THIN_BORDER
    current_row += 1

    neither_start_row = current_row
    for _, n_row in neither.iterrows():
        for col_idx, col_name in enumerate(NEITHER_COLS, 1):
            val = n_row.get(col_name)
            cell = ws.cell(row=current_row, column=col_idx, value=val if pd.notna(val) else '')
            cell.font = NEITHER_FONT
            cell.fill = NEITHER_FILL
            cell.border = THIN_BORDER
            fmt = NUMBER_FORMATS.get(col_name)
            if fmt and pd.notna(val):
                cell.number_format = fmt
        current_row += 1
    neither_end_row = current_row - 1

    # Group/collapse neither rows using Excel outline
    if neither_end_row >= neither_start_row:
        ws.row_dimensions.group(neither_start_row, neither_end_row, outline_level=1)

    # ── Conditional formatting: highlight best in each metric column ──
    if qual_end_row >= qual_start_row:
        # Columns to highlight (1-based indices in DETAIL_COLS)
        # CAGR_pct=3, Vol_pct=4 (lower is better), Sharpe=5, Sortino=6,
        # Calmar=7, MaxDD_pct=8 (higher/less negative is better),
        # End_Value=9, Pct_Cash_Time=13 (lower is better),
        # Trades_Per_Year=14 (lower is better), Avg_Leverage=16
        highlight_cols_green = {
            3: 'max',   # CAGR — higher is better
            5: 'max',   # Sharpe — higher is better
            6: 'max',   # Sortino — higher is better
            7: 'max',   # Calmar — higher is better
            8: 'max',   # MaxDD — less negative is better
            9: 'max',   # End_Value — higher is better
        }
        highlight_cols_red = {
            4: 'min',   # Vol — lower is better
            13: 'min',  # Pct_Cash_Time — lower is better
            14: 'min',  # Trades_Per_Year — lower is better
        }

        for col_idx, mode in highlight_cols_green.items():
            col_letter = get_column_letter(col_idx)
            rng = f'{col_letter}{qual_start_row}:{col_letter}{qual_end_row}'
            if mode == 'max':
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f'{col_letter}{qual_start_row}=MAX(${col_letter}${qual_start_row}:${col_letter}${qual_end_row})'],
                        fill=PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
                        font=Font(bold=True, color='FFFFFF')
                    )
                )
            else:
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f'{col_letter}{qual_start_row}=MIN(${col_letter}${qual_start_row}:${col_letter}${qual_end_row})'],
                        fill=PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
                        font=Font(bold=True, color='FFFFFF')
                    )
                )

        for col_idx, mode in highlight_cols_red.items():
            col_letter = get_column_letter(col_idx)
            rng = f'{col_letter}{qual_start_row}:{col_letter}{qual_end_row}'
            if mode == 'min':
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f'{col_letter}{qual_start_row}=MIN(${col_letter}${qual_start_row}:${col_letter}${qual_end_row})'],
                        fill=PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid'),
                        font=Font(bold=True, color='FFFFFF')
                    )
                )
            else:
                ws.conditional_formatting.add(
                    rng,
                    FormulaRule(
                        formula=[f'{col_letter}{qual_start_row}=MAX(${col_letter}${qual_start_row}:${col_letter}${qual_end_row})'],
                        fill=PatternFill(start_color='FF4444', end_color='FF4444', fill_type='solid'),
                        font=Font(bold=True, color='FFFFFF')
                    )
                )

    # Freeze panes (row 5, after section header + trading cost + B&H header)
    ws.freeze_panes = 'A5'

    # Auto-fit columns
    auto_fit_columns(ws)

    print(f"  Sheet '{sheet_name}': {len(qualifying)} {classification}, "
          f"{len(neither)} neither", flush=True)
    return ws


def build_stillwater_sheet(wb, df, asset_key):
    """Build a Stillwater detail sheet for one asset (e.g. Nasdaq 100).

    Stillwater = best Water/Octane-style variants accepting CAGR up to
    STILLWATER_CAGR_TOL percentage points below Buy & Hold 1x. See
    classify_stillwater() for the exact rule.
    """
    asset_label = ASSET_LABELS[asset_key]
    sheet_name = f'{asset_label} Stillwater'[:31]
    ws = wb.create_sheet(title=sheet_name)

    asset_df = df[df['Asset'] == asset_key].copy()
    bh_rows = asset_df[asset_df['Strategy'].apply(is_bh_row)]
    strategies_df = asset_df[~asset_df['Strategy'].apply(is_bh_row)].copy()

    bh1 = bh_rows[bh_rows['Strategy'] == 'Buy & Hold 1x']
    bh_metrics = None
    if len(bh1) > 0:
        bh_metrics = {k: float(bh1.iloc[0][k]) for k in
                      ['Sharpe', 'Calmar', 'MaxDD_pct', 'CAGR_pct', 'Sortino', 'Vol_pct']}

    strategies_df['SW'] = strategies_df.apply(lambda r: classify_stillwater(r, bh_metrics), axis=1)
    water = strategies_df[strategies_df['SW'] == 'Stillwater-Water'].sort_values('Sharpe', ascending=False)
    octane = strategies_df[strategies_df['SW'] == 'Stillwater-Octane'].sort_values('Calmar', ascending=False)
    neither = strategies_df[~strategies_df['SW'].isin(['Stillwater-Water', 'Stillwater-Octane'])
                            ].sort_values('MaxDD_pct', ascending=True)

    def write_header(r):
        for col_idx, col_name in enumerate(DETAIL_COLS, 1):
            cell = ws.cell(row=r, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER
        return r + 1

    def write_row(r, data, font, fill):
        for col_idx, col_name in enumerate(DETAIL_COLS, 1):
            val = data.get(col_name)
            cell = ws.cell(row=r, column=col_idx, value=val if pd.notna(val) else '')
            cell.font = font
            cell.fill = fill
            cell.border = THIN_BORDER
            fmt = NUMBER_FORMATS.get(col_name)
            if fmt and pd.notna(val):
                cell.number_format = fmt
        return r + 1

    cr = 1
    # ── Title ──
    cell = ws.cell(row=cr, column=1, value=f'{asset_label} — Stillwater Strategies')
    cell.font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(DETAIL_COLS))
    cr += 1
    # ── Rule subtitle ──
    if bh_metrics:
        bh_cagr = bh_metrics['CAGR_pct']
        sub = (f"Water/Octane risk criteria, but CAGR may be up to {STILLWATER_CAGR_TOL:.0f}pp "
               f"below Buy & Hold 1x (CAGR >= {bh_cagr - STILLWATER_CAGR_TOL:.2f}% vs B&H {bh_cagr:.2f}%). "
               f"For assets where strict Water/Octane is unreachable. "
               f"Trading cost: {TRADING_COST_LABELS.get(asset_key, '')} from mid-price.")
    else:
        sub = "Water/Octane risk criteria with CAGR relaxed by up to 2pp vs Buy & Hold 1x."
    cell = ws.cell(row=cr, column=1, value=sub)
    cell.font = Font(name='Calibri', size=9, italic=True, color='808080')
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(DETAIL_COLS))
    cr += 1

    # ── Benchmarks ──
    cell = ws.cell(row=cr, column=1, value='Benchmarks (Buy & Hold):')
    cell.font = Font(name='Calibri', size=11, bold=True, italic=True)
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(DETAIL_COLS))
    cr += 1
    cr = write_header(cr)
    for _, bh_row in bh_rows.sort_values('Leverage_Max', ascending=True).iterrows():
        cr = write_row(cr, bh_row, BH_FONT, BH_FILL)

    # ── Qualifying subsections ──
    for title, frame in [('Stillwater — Water-style', water), ('Stillwater — Octane-style', octane)]:
        cr += 1
        cell = ws.cell(row=cr, column=1, value=f'{title} ({len(frame)}):')
        cell.font = Font(name='Calibri', size=11, bold=True, color='375623')
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(DETAIL_COLS))
        cr += 1
        cr = write_header(cr)
        for _, q_row in frame.iterrows():
            cr = write_row(cr, q_row, NORMAL_FONT, WATER_FILL)

    # ── Did-not-qualify (collapsed) ──
    cr += 1
    cell = ws.cell(row=cr, column=1,
                   value=f'Did Not Qualify ({len(neither)}) — collapsed (Strategy + MaxDD only):')
    cell.font = Font(name='Calibri', size=11, bold=True, color='808080')
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(NEITHER_COLS))
    cr += 1
    for col_idx, col_name in enumerate(NEITHER_COLS, 1):
        cell = ws.cell(row=cr, column=col_idx, value=col_name)
        cell.font = Font(name='Calibri', size=11, bold=True, color='595959')
        cell.fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
        cell.border = THIN_BORDER
    cr += 1
    n_start = cr
    for _, n_row in neither.iterrows():
        for col_idx, col_name in enumerate(NEITHER_COLS, 1):
            val = n_row.get(col_name)
            cell = ws.cell(row=cr, column=col_idx, value=val if pd.notna(val) else '')
            cell.font = NEITHER_FONT
            cell.fill = NEITHER_FILL
            cell.border = THIN_BORDER
            fmt = NUMBER_FORMATS.get(col_name)
            if fmt and pd.notna(val):
                cell.number_format = fmt
        cr += 1
    if cr - 1 >= n_start:
        ws.row_dimensions.group(n_start, cr - 1, outline_level=1)

    ws.freeze_panes = 'A5'
    auto_fit_columns(ws)
    print(f"  Sheet '{sheet_name}': {len(water)} Water-style, {len(octane)} Octane-style "
          f"Stillwater, {len(neither)} did not qualify", flush=True)
    return ws


def auto_fit_columns(ws, min_width=8, max_width=40):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # Estimate width: count characters, adjust for number formats
                val_str = str(cell.value)
                # Rough adjustment: each char ~1.1 units, add padding
                cell_len = len(val_str) * 1.1 + 2
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = max(min_width, min(max_len, max_width))
        ws.column_dimensions[col_letter].width = adjusted


def build_lqq3_sheet(wb):
    """LQQ3 dedicated backtest sheet: signal computed ON LQQ3 vs ON the underlying Nasdaq,
    over real LQQ3 (2012+) and a synthetic 3x model (1990+). Flat table, no Water/Octane.
    Reads output/lqq3_dedicated/results.csv (from research/analyze_lqq3_signal_source.py)."""
    csv_path = ROOT / 'output' / 'lqq3_dedicated' / 'results.csv'
    if not csv_path.exists():
        print("  (skip LQQ3 sheet — run research/analyze_lqq3_signal_source.py first)", flush=True)
        return
    ldf = pd.read_csv(csv_path)
    ws = wb.create_sheet(title='LQQ3', index=1)

    cols = ['Strategy', 'Signal_Source', 'Data', 'Start', 'End', 'Trading_Days',
            'CAGR_pct', 'Vol_pct', 'Sharpe', 'MaxDD_pct', 'Calmar', 'End_Value', 'Pct_Cash']
    headers = ['Strategy', 'Signal source', 'Data', 'Start', 'End', 'Days',
               'CAGR', 'Vol', 'Sharpe', 'MaxDD', 'Calmar', 'End $', '% Cash']
    fmts = {'CAGR_pct': '0.00"%"', 'Vol_pct': '0.00"%"', 'MaxDD_pct': '0.00"%"',
            'Pct_Cash': '0.0"%"', 'Sharpe': '0.000', 'Calmar': '0.000',
            'End_Value': '$#,##0', 'Trading_Days': '#,##0'}
    widths = {1: 32, 2: 24, 3: 20, 4: 12, 5: 12, 6: 7, 7: 9, 8: 8, 9: 8, 10: 9, 11: 8, 12: 12, 13: 8}
    for ci, w in widths.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    side = Side(style='thin', color='D9D9D9')
    thin = Border(left=side, right=side, top=side, bottom=side)
    hdr_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    base_font = Font(name='Calibri', size=10)
    bh_font = Font(name='Calibri', size=10, italic=True, color='808080')
    under_fill = PatternFill('solid', fgColor='E8F0FE')  # tint signal-on-underlying rows

    cr = 1
    c = ws.cell(cr, 1, 'LQQ3.L (3x Nasdaq ETP) — signal on LQQ3 vs on the underlying Nasdaq 100')
    c.font = Font(name='Calibri', size=14, bold=True, color='1F4E79')
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(cols)); cr += 1
    sub = ("Each signal strategy run two ways — signal from the held 3x instrument's own price vs from the "
           "underlying ^NDX index — over real LQQ3 (listing 2012-12-13) and a synthetic 3x daily-reset model on "
           "^NDX (1990+). One flat table, no Water/Octane. Engine: 1-day signal lag, 0.10% cost, $100 + $10/yr "
           "inflow. Shaded rows = signal on the underlying Nasdaq.")
    c = ws.cell(cr, 1, sub); c.font = Font(name='Calibri', size=9, italic=True, color='808080')
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(cols))
    ws.row_dimensions[cr].height = 42; cr += 2

    for regime in ['Real LQQ3 (2012+)', 'Synthetic 3x (1990+)']:
        sub_rows = ldf[ldf['Data'] == regime]
        if sub_rows.empty:
            continue
        c = ws.cell(cr, 1, regime)
        c.font = Font(name='Calibri', size=11, bold=True, color='1F4E79')
        ws.merge_cells(start_row=cr, start_column=1, end_row=cr, end_column=len(cols)); cr += 1
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(cr, ci, h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center'); cell.border = thin
        cr += 1
        for _, r in sub_rows.iterrows():
            is_under = str(r['Signal_Source']).startswith('Nasdaq')
            is_bh = str(r['Strategy']).startswith('Buy & hold')
            for ci, col in enumerate(cols, 1):
                val = r[col]
                cell = ws.cell(cr, ci, None if pd.isna(val) else val)
                cell.font = bh_font if is_bh else base_font
                cell.border = thin
                if is_under:
                    cell.fill = under_fill
                if col in fmts and pd.notna(val):
                    cell.number_format = fmts[col]
            cr += 1
        cr += 1
    ws.freeze_panes = 'A4'


def main():
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    print("=" * 70, flush=True)
    print("BUILDING STRATEGY RESULTS EXCEL WORKBOOK", flush=True)
    print("=" * 70, flush=True)

    # ── Step 1: Load data ──
    print("\n[1/5] Loading CSV data...", flush=True)
    df = load_all_data()
    print(f"  Total rows: {len(df)}", flush=True)
    print(f"  Assets: {sorted(df['Asset'].unique())}", flush=True)

    # ── Step 2: Classify ──
    print("\n[2/5] Classifying strategies (Water / Octane / Neither)...", flush=True)
    df = classify_all(df)

    # Print classification summary
    for asset_key in ASSET_ORDER:
        asset_df = df[df['Asset'] == asset_key]
        strats = asset_df[~asset_df['Strategy'].apply(is_bh_row)]
        w = (strats['Classification'] == 'Water').sum()
        o = (strats['Classification'] == 'Octane').sum()
        n = (strats['Classification'] == 'Neither').sum()
        print(f"  {ASSET_LABELS[asset_key]:20s}: {len(strats):2d} strategies -> "
              f"Water={w}, Octane={o}, Neither={n}", flush=True)

    # ── Step 3: Create workbook ──
    print("\n[3/5] Creating Excel workbook...", flush=True)
    wb = openpyxl.Workbook()

    # ── Step 4: Build sheets ──
    print("\n[4/5] Building sheets...", flush=True)

    # Definitions sheet
    print("  Building Definitions sheet...", flush=True)
    build_definitions_sheet(wb)

    # LQQ3 dedicated sheet (signal on LQQ3 vs underlying Nasdaq; real vs synthetic)
    print("  Building LQQ3 sheet...", flush=True)
    build_lqq3_sheet(wb)

    # Summary sheet
    print("  Building Summary sheet...", flush=True)
    build_summary_sheet(wb, df)

    # Detail sheets (Water + Octane per asset; Nasdaq also gets a Stillwater sheet)
    for asset_key in ASSET_ORDER:
        for classification in ['Water', 'Octane']:
            build_detail_sheet(wb, df, asset_key, classification)
        if asset_key == 'ndx':
            build_stillwater_sheet(wb, df, asset_key)

    # Remove the default empty 'Sheet' that openpyxl creates with the workbook
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # ── Step 5: Save ──
    print("\n[5/5] Saving workbook...", flush=True)
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"  Saved to: {EXCEL_PATH}", flush=True)
    print(f"  File size: {EXCEL_PATH.stat().st_size:,} bytes", flush=True)
    print(f"  Sheets: {len(wb.sheetnames)}", flush=True)

    # ── Final report ──
    print("\n" + "=" * 70, flush=True)
    print("BUILD COMPLETE", flush=True)
    print("=" * 70, flush=True)

    # Summary statistics
    strategies_df = df[~df['Strategy'].apply(is_bh_row)]
    total_strategies = len(strategies_df)
    total_water = (strategies_df['Classification'] == 'Water').sum()
    total_octane = (strategies_df['Classification'] == 'Octane').sum()
    total_neither = (strategies_df['Classification'] == 'Neither').sum()

    print(f"\nOverall Statistics:")
    print(f"  Total strategies (excl. B&H): {total_strategies}")
    print(f"  Water:  {total_water}")
    print(f"  Octane: {total_octane}")
    print(f"  Neither: {total_neither}")
    print(f"  Sheets: {len(wb.sheetnames)} (1 Summary + 20 detail)")

    # Top Water strategies overall
    water_df = strategies_df[strategies_df['Classification'] == 'Water']
    if len(water_df) > 0:
        print(f"\nTop 5 Water Strategies (by Sharpe):")
        top_water = water_df.nlargest(5, 'Sharpe')
        for _, r in top_water.iterrows():
            print(f"  {r['Asset']:10s} | {r['Strategy']:45s} | "
                  f"Sharpe={r['Sharpe']:.3f} | CAGR={r['CAGR_pct']:.2f}% | "
                  f"MaxDD={r['MaxDD_pct']:.2f}%")

    # Top Octane strategies overall
    octane_df = strategies_df[strategies_df['Classification'] == 'Octane']
    if len(octane_df) > 0:
        print(f"\nTop 5 Octane Strategies (by Calmar):")
        top_octane = octane_df.nlargest(5, 'Calmar')
        for _, r in top_octane.iterrows():
            print(f"  {r['Asset']:10s} | {r['Strategy']:45s} | "
                  f"Calmar={r['Calmar']:.3f} | CAGR={r['CAGR_pct']:.2f}% | "
                  f"MaxDD={r['MaxDD_pct']:.2f}%")

    print(f"\nExcel location: {EXCEL_PATH}")
    print("Done.", flush=True)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f"ERROR: {e}", flush=True)
        traceback.print_exc()
        sys.exit(1)
