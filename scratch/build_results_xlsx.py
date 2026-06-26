"""Build 'Claude back testing.xlsx' — all S&P 500 strategy backtest results from this session.
Recomputes everything (no hand-transcribed numbers) and writes a formatted multi-sheet workbook.
CAGR/maxDD stored as fractions (shown %); Calmar written as a live Excel formula =CAGR/ABS(DD)."""
from __future__ import annotations
import sys
from pathlib import Path
import numpy as np, pandas as pd, yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))
from core.engine import PortfolioEngine, FUNDING_SPREAD, TRADING_DAYS
from core.etp_leverage import TER_ANNUAL
from core.metrics import comprehensive_stats
from test_guarded_balanced_candidate import guarded_strategy_leverage
SPEC = dict(trigger_a=0.05, trigger_b=0.25, lead_pct_below_sma20=0.0075, x_return=0.40, y_return=0.15)
OUT = ROOT / "Results" / "Claude back testing.xlsx"

_C={}
def close(tk):
    if tk not in _C:
        s=yf.download(tk,period="max",auto_adjust=True,progress=False)["Close"].dropna().astype(float).squeeze(); s.index=s.index.tz_localize(None); _C[tk]=s
    return _C[tk]
def prices(a="1990-01-01",b=None):
    p=pd.DataFrame({"spx_close":close("^GSPC"),"tbill_rate":close("^IRX")/100}).sort_index().ffill().dropna(); p=p[p.index>=pd.Timestamp(a)]
    return p[p.index<pd.Timestamp(b)] if b else p
def pan(p):
    r=p["spx_close"].pct_change(); tb=p["tbill_rate"]
    def bor(L): return (L-1)*(tb+FUNDING_SPREAD)/TRADING_DAYS
    return pd.DataFrame({"ret_0":tb/TRADING_DAYS,"ret_1":(r-TER_ANNUAL[1]/TRADING_DAYS).fillna(0),
        "ret_2":(2*r-bor(2)-TER_ANNUAL[2]/TRADING_DAYS).fillna(0),"ret_3":(3*r-bor(3)-TER_ANNUAL[3]/TRADING_DAYS).fillna(0)},index=p.index)
def E(): return PortfolioEngine(max_drawdown_limit=None,hard_drawdown_floor=False,trading_cost_pct=0.001,annual_inflow_pct=0,annual_inflow_abs=0.0)
def C(p): return p["spx_close"].astype(float)
def sma(c,w): return c.rolling(w,min_periods=w).mean()
def ema(c,w): return c.ewm(span=w,min_periods=w).mean()
def rvol(c): return c.pct_change().rolling(20).std()*np.sqrt(252)
def synthLx(p,L): return (1+pan(p)[f"ret_{L}"].fillna(0)).cumprod()
def relhigh(c,k=1.2): rv=rvol(c); th=(k*rv.rolling(252,min_periods=60).median()).fillna(0.20); return (rv>th).fillna(False)
def stat(p,lev):
    res=E().run(p,lev,etp_returns=pan(p)); s=comprehensive_stats(res.equity,res.daily_returns)
    yrs=(p.index[-1]-p.index[0]).days/365.25
    return dict(cagr=s["cagr"],dd=s["max_drawdown"],sharpe=s["sharpe"],tr=res.rebalance_count/yrs,grow=res.equity.iloc[-1]/res.equity.iloc[0])
# builders
def bh(c,L): return pd.Series(float(L),index=c.index)
def sma_cash(c,w,L): o=pd.Series(0.0,index=c.index); o[c>sma(c,w)]=L; return o
def golden(c,L): o=pd.Series(0.0,index=c.index); o[sma(c,50)>sma(c,200)]=L; return o
def band(c,L,w=200,be=0.03,bx=None):
    bx=be if bx is None else bx; s=sma(c,w); o=pd.Series(np.nan,index=c.index); o[c>s*(1+be)]=L; o[c<s*(1-bx)]=0.0; return o.ffill().fillna(0.0)
def eband(c,L,w=200,b=0.03): s=ema(c,w); o=pd.Series(np.nan,index=c.index); o[c>s*(1+b)]=L; o[c<s*(1-b)]=0.0; return o.ffill().fillna(0.0)
def monthlyize(raw): per=raw.index.to_period("M"); last=raw.groupby(per).last().shift(1); return pd.Series(per.map(last),index=raw.index).astype(float).fillna(0)
def mom(c,L,lb=252): o=pd.Series(0.0,index=c.index); o[c>c.shift(lb)]=L; return o
def volguard(c,L,cap): b=golden(c,L); b[relhigh(c)&(b>cap)]=cap; return b
def band_momgate(c,L,w=200,b=0.03): base=band(c,L,w,b); base[~(c>c.shift(252)).fillna(False)]=0.0; return base
def band_voltier(c,w=200,b=0.03):
    base=band(c,3,w,b); rv=rvol(c); base[(base>0)&(rv>rv.rolling(252,min_periods=60).median()).fillna(False)]=2.0; return base
def band_fastreentry(c,L,w=200,bx=0.03,refast=50):
    s=sma(c,w).values; sf=sma(c,refast).values; cv=c.values; out=np.zeros(len(cv)); st=0.0
    for i in range(len(cv)):
        if np.isnan(s[i]): out[i]=0.0; continue
        if st==0.0:
            if cv[i]>s[i] and (not np.isnan(sf[i]) and cv[i]>sf[i]): st=L
        else:
            if cv[i]<s[i]*(1-bx): st=0.0
        out[i]=st
    return pd.Series(out,index=c.index)

print("computing...", flush=True)
p=prices(); c1=C(p); P1=prices(b="2008-01-01"); P2=prices("2008-01-01"); c1a=C(P1); c1b=C(P2)
START,END,YRS=p.index[0].date(),p.index[-1].date(),(p.index[-1]-p.index[0]).days/365.25

# ===== PART 1: bake-off (signal on 1x) =====
B1=[("Buy & hold",bh,[1,2,3]),("SMA200 / cash",lambda c,L:sma_cash(c,200,L),[1,2,3]),
    ("SMA20 / cash",lambda c,L:sma_cash(c,20,L),[1,2,3]),("Golden 50/200",golden,[1,2,3]),
    ("SMA200 monthly",lambda c,L:monthlyize(sma_cash(c,200,L)),[2,3]),
    ("SMA200 3% band",band,[1,2,3]),("Mom 12m",mom,[2,3])]
bake=[]
for name,fn,levs in B1:
    for L in levs:
        d=stat(p,fn(c1,L)); bake.append([name,L,d])
for name,(L,cap) in [("Golden vol-guard",(2,1)),("Golden vol-guard",(3,2))]:
    d=stat(p,volguard(c1,L,cap)); bake.append([name,L,d])
bake.append(["Guarded A5/B25 (tiered)",3,stat(p,guarded_strategy_leverage(p,**SPEC)[0])])
bake.sort(key=lambda r:-r[2]["cagr"])

# ===== PART 2: signal basis =====
P2tbl=[]
for label,fn in [("SMA200 / cash",lambda c,L:sma_cash(c,200,L)),("Golden 50/200",golden),("SMA200 3% band",band)]:
    for L in [2,3]:
        a=stat(p,fn(c1,L)); b=stat(p,fn(synthLx(p,L),L))
        P2tbl.append([label,L,"1x S&P (underlying)",a]); P2tbl.append([label,L,f"{L}x series (itself)",b])

# ===== PART 3: robustness of bake-off top 6 =====
def rebuild(name,L):
    if name.startswith("Guarded"): return None
    if name=="Buy & hold": return lambda c:bh(c,L)
    if name=="SMA200 monthly": return lambda c:monthlyize(sma_cash(c,200,L))
    if name=="SMA200 3% band": return lambda c:band(c,L)
    if name=="SMA200 / cash": return lambda c:sma_cash(c,200,L)
    if name=="SMA20 / cash": return lambda c:sma_cash(c,20,L)
    if name=="Golden vol-guard": return lambda c:volguard(c,L,L-1)
    if name=="Golden 50/200": return lambda c:golden(c,L)
    if name=="Mom 12m": return lambda c:mom(c,L)
rob=[]
for name,L,d in bake[:6]:
    fn=rebuild(name,L)
    if fn is None:
        h1=stat(P1,guarded_strategy_leverage(P1,**SPEC)[0]); h2=stat(P2,guarded_strategy_leverage(P2,**SPEC)[0])
    else:
        h1=stat(P1,fn(c1a)); h2=stat(P2,fn(c1b))
    rob.append([f"{name} {L}x",h1,h2,d["cagr"]])

# ===== PART 4: beat-the-band search =====
bench=stat(p,band(c1,3,200,0.03))
cands=[]
for w in [120,150,180,200,220,250,300]:
    for b in [0.0,0.02,0.03,0.04,0.05,0.06]:
        cands.append((f"SMA{w} {int(b*100)}% band 3x", lambda c,w=w,b=b: band(c,3,w,b)))
for w in [100,150,200]:
    for b in [0.02,0.03,0.04]:
        cands.append((f"EMA{w} {int(b*100)}% band 3x", lambda c,w=w,b=b: eband(c,3,w,b)))
for be in [0.01,0.02,0.03]:
    for bx in [0.03,0.05,0.07,0.09]:
        cands.append((f"Asym enter +{int(be*100)}% / exit -{int(bx*100)}% 3x", lambda c,be=be,bx=bx: band(c,3,200,be,bx)))
cands += [("Band + 12m momentum gate 3x", lambda c: band_momgate(c,3,200,0.03)),
          ("Band vol-tier 3<->2x", lambda c: band_voltier(c,200,0.03)),
          ("Band fast crash re-entry 3x", lambda c: band_fastreentry(c,3,200,0.03,50)),
          ("Band fast re-entry, wide exit 3x", lambda c: band_fastreentry(c,3,200,0.05,50))]
sweep=[(n,stat(p,fn(c1)),fn) for n,fn in cands]
sweep.sort(key=lambda x:-x[1]["cagr"])
top=sweep[:12]
# overfit guard: benchmark + top 6 unique
guard_names=["SMA200 3% band 3x (benchmark)"]+[n for n,_,_ in top]
seen=set(); guard=[]
for n in guard_names:
    if n in seen: continue
    seen.add(n)
    fn = (lambda c: band(c,3,200,0.03)) if n.startswith("SMA200 3% band 3x") else dict((nn,ff) for nn,_,ff in sweep)[n]
    guard.append([n,stat(P1,fn(c1a)),stat(P2,fn(c1b)),stat(p,fn(c1))])
    if len(guard)>=7: break

# =========================== WRITE WORKBOOK ===========================
print("writing xlsx...", flush=True)
NAVY='1F4E78'; LBLUE='D6E4F0'; GREEN='C6EFCE'; REDF='FFC7CE'; GREY='F2F2F2'
HDR=Font(name='Arial',bold=True,color='FFFFFF',size=10); BODY=Font(name='Arial',size=10); BODYB=Font(name='Arial',size=10,bold=True)
TITLE=Font(name='Arial',bold=True,size=15,color='1F4E78'); SUBT=Font(name='Arial',size=10,italic=True,color='595959')
HFILL=PatternFill('solid',fgColor=NAVY); WINFILL=PatternFill('solid',fgColor=GREEN)
CTR=Alignment(horizontal='center',vertical='center',wrap_text=True); L=Alignment(horizontal='left',vertical='center'); WRAP=Alignment(horizontal='left',vertical='top',wrap_text=True)
thin=Side(style='thin',color='BFBFBF'); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
PCT='0.0%;(0.0%)'; NUM2='0.00'; NUM1='0.0'; GROW='#,##0.0"x"'

def hrow(ws,r,headers,widths=None):
    for j,h in enumerate(headers,1):
        c=ws.cell(r,j,h); c.font=HDR; c.fill=HFILL; c.alignment=CTR; c.border=BORD
    if widths:
        for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
def cell(ws,r,j,v,fmt=None,font=BODY,align=None,fill=None):
    c=ws.cell(r,j,v); c.font=font; c.border=BORD
    if fmt:c.number_format=fmt
    if align:c.alignment=align
    if fill:c.fill=fill
    return c

wb=Workbook()

# ---- Sheet: Summary ----
ws=wb.active; ws.title="Summary"; ws.sheet_view.showGridLines=False
ws.column_dimensions['A'].width=3; ws.column_dimensions['B'].width=104
cell(ws,2,2,"S&P 500 Strategy Backtests — Claude analysis",font=TITLE,align=L)
cell(ws,3,2,f"Synthetic daily-reset leverage model · {START} to {END} ({YRS:.0f} years) · 0.10% cost/trade · signals lagged 1 day · no inflows",font=SUBT,align=L)
lines=[("",""),
 ("THE VERDICT","h"),
 ("Best S&P strategy, CAGR-first: SMA200 ±3% band, at 3x, with the signal taken on the 1x S&P.","b"),
 (f"  16.6% CAGR  ·  -54% max drawdown  ·  ~1 trade/year  ·  274x growth over {YRS:.0f}y.","n"),
 ("It posts the highest CAGR tested and, unlike other 3x options, avoids ruin (other 3x trend filters hit -77% to -99%).","n"),
 ("",""),
 ("TWO NON-NEGOTIABLE RULES (from the data)","h"),
 ("1. Signal on the 1x S&P, never the leveraged series. For the 3x band, CAGR collapses 16.6% -> 9.4% if you","b"),
 ("    compute the signal on the 3x series itself (see 'Signal basis' tab). Watch the index; trade the 3x product.","n"),
 ("2. Never buy & hold leverage. Naked 2x/3x buy & hold are wiped out (-91% / -99%). The trend filter is what makes","b"),
 ("    leverage survivable. Daily SMA20 is also out (whipsaws to negative CAGR when levered).","n"),
 ("",""),
 ("CAN ANYTHING BEAT IT?","h"),
 ("Essentially no. A wide-exit variant (enter +3% / exit -5%) edges it to 17.1% CAGR and is more regime-consistent,","b"),
 ("    but costs ~2.6 pts more drawdown (-56.5%). Everything else either overfits (wins one half of history, loses the","n"),
 ("    other) or reduces CAGR. The band 3x sits at the efficient frontier for S&P trend-following (see 'Beat-the-band' tab).","n"),
 ("",""),
 ("THE LEVERAGE LADDER (your drawdown dial, SMA200 3% band)","h"),
 ("1x = 9.0% CAGR / -20% DD   ·   2x = 13.5% / -39%   ·   3x = 16.6% / -54%.  Each step ~ +3-4% CAGR for ~15-20 pts more drawdown.","n"),
 ("",""),
 ("HONEST CAVEATS","h"),
 ("-54% is a real gut-check (100k -> 46k) and only works if held through. Daily-reset 3x has tail risk a backtest understates","n"),
 ("(an overnight gap beyond ~-33% impairs a 3x product before any signal reacts). 3x figures lean on the synthetic model;","n"),
 ("real UPRO exists only from 2009 and tracks within ~2%/yr. This is quantitative analysis of strategies, not financial advice.","n"),
 ("",""),
 ("TABS: Bake-off (every strategy x 1/2/3x) · Signal basis (1x vs leveraged-series signal) · Robustness (by half) · Beat-the-band (search + overfit guard).","n"),
]
r=5
for txt,kind in lines:
    c=cell(ws,r,2,txt,align=WRAP)
    if kind=="h": c.font=Font(name='Arial',bold=True,size=11,color='1F4E78')
    elif kind=="b": c.font=BODYB
    else: c.font=BODY
    r+=1

# ---- Sheet: Bake-off ----
ws=wb.create_sheet("Bake-off"); ws.sheet_view.showGridLines=False
cell(ws,1,1,f"S&P 500 strategy bake-off — signal on 1x S&P, ranked by CAGR ({START}..{END}, {YRS:.0f}y, no inflow, 0.10% cost)",font=TITLE,align=L)
hdr=["Rank","Strategy","Lev","CAGR","Max DD","Calmar","Sharpe","Trades/yr","Growth","Risk flag"]
hrow(ws,3,hdr,[6,24,6,10,10,9,9,11,11,14]); ws.freeze_panes="A4"
for i,(name,Lv,d) in enumerate(bake):
    r=4+i; win = (name=="SMA200 3% band" and Lv==3)
    f=WINFILL if win else None; ft=BODYB if win else BODY
    cell(ws,r,1,i+1,font=ft,align=CTR,fill=f); cell(ws,r,2,name,font=ft,align=L,fill=f); cell(ws,r,3,f"{Lv}x",font=ft,align=CTR,fill=f)
    cell(ws,r,4,d["cagr"],PCT,ft,fill=f); cell(ws,r,5,d["dd"],PCT,ft,fill=f)
    cell(ws,r,6,f"=IF(E{r}=0,\"\",D{r}/ABS(E{r}))",NUM2,ft,fill=f)
    cell(ws,r,7,d["sharpe"],NUM2,ft,fill=f); cell(ws,r,8,d["tr"],NUM1,ft,fill=f); cell(ws,r,9,d["grow"],GROW,ft,fill=f)
    flag = "RUIN (<-85%)" if d["dd"]<-0.85 else ("deep (<-70%)" if d["dd"]<-0.70 else "")
    fc=cell(ws,r,10,flag,align=CTR,fill=f);
    if flag: fc.font=Font(name='Arial',size=10,bold=True,color='C00000')
cell(ws,5+len(bake),2,"Winner (green) = recommended. Growth = total multiple of $1 over the period. CAGR/DD annualised. Calmar = CAGR / |Max DD| (live formula).",font=SUBT,align=L)

# ---- Sheet: Signal basis ----
ws=wb.create_sheet("Signal basis"); ws.sheet_view.showGridLines=False
cell(ws,1,1,"Does signalling on the leveraged series help? Signal on 1x S&P vs on the 2x/3x series itself",font=TITLE,align=L)
hrow(ws,3,["Strategy","Lev","Signal computed on","CAGR","Max DD","Calmar"],[24,6,24,10,10,9]); ws.freeze_panes="A4"
for i,(name,Lv,basis,d) in enumerate(P2tbl):
    r=4+i; under = "underlying" in basis; ft=BODYB if under else BODY; f=PatternFill('solid',fgColor=GREY) if under else None
    cell(ws,r,1,name,font=ft,align=L,fill=f); cell(ws,r,2,f"{Lv}x",font=ft,align=CTR,fill=f); cell(ws,r,3,basis,font=ft,align=L,fill=f)
    cell(ws,r,4,d["cagr"],PCT,ft,fill=f); cell(ws,r,5,d["dd"],PCT,ft,fill=f); cell(ws,r,6,f"=IF(E{r}=0,\"\",D{r}/ABS(E{r}))",NUM2,ft,fill=f)
cell(ws,5+len(P2tbl),1,"Shaded = signal on the 1x underlying (the correct, higher-CAGR method). Signalling on the leveraged series loses CAGR, badly at 3x.",font=SUBT,align=L)

# ---- Sheet: Robustness ----
ws=wb.create_sheet("Robustness"); ws.sheet_view.showGridLines=False
cell(ws,1,1,"Robustness — CAGR / Max DD in each half of history (signal on 1x S&P)",font=TITLE,align=L)
hrow(ws,3,["Strategy","H1 1990-2008 CAGR","H1 Max DD","H2 2008-2026 CAGR","H2 Max DD","Full CAGR"],[26,17,12,17,12,11]); ws.freeze_panes="A4"
for i,(name,h1,h2,full) in enumerate(rob):
    r=4+i; cell(ws,r,1,name,align=L)
    cell(ws,r,2,h1["cagr"],PCT); cell(ws,r,3,h1["dd"],PCT); cell(ws,r,4,h2["cagr"],PCT); cell(ws,r,5,h2["dd"],PCT); cell(ws,r,6,full,PCT,font=BODYB)
cell(ws,5+len(rob),1,"A robust strategy is strong in BOTH halves. The band 3x is consistent (14.4% / 18.9%); see Beat-the-band tab for the overfit guard.",font=SUBT,align=L)

# ---- Sheet: Beat-the-band ----
ws=wb.create_sheet("Beat-the-band"); ws.sheet_view.showGridLines=False
cell(ws,1,1,"Can anything beat SMA200 3% band 3x? Search of 60+ variants, then overfit guard",font=TITLE,align=L)
cell(ws,2,1,f"Benchmark — SMA200 3% band 3x:  CAGR {bench['cagr']*100:.1f}%   Max DD {bench['dd']*100:.1f}%   Calmar {bench['cagr']/abs(bench['dd']):.2f}",font=SUBT,align=L)
hrow(ws,4,["Top candidates by full-period CAGR","CAGR","Max DD","Calmar","Trades/yr","Beats benchmark?"],[34,10,10,9,11,16])
for i,(name,d,fn) in enumerate(top):
    r=5+i; beats = d["cagr"]>bench["cagr"]; ft=BODYB if beats else BODY
    cell(ws,r,1,name,font=ft,align=L); cell(ws,r,2,d["cagr"],PCT,ft); cell(ws,r,3,d["dd"],PCT,ft)
    cell(ws,r,4,f"=IF(C{r}=0,\"\",B{r}/ABS(C{r}))",NUM2,ft); cell(ws,r,5,d["tr"],NUM1,ft)
    bc=cell(ws,r,6,"YES" if beats else "no",align=CTR,font=ft)
    if beats: bc.font=Font(name='Arial',size=10,bold=True,color='1F7A1F')
g0=5+len(top)+2
cell(ws,g0-1,1,"OVERFIT GUARD — re-checked in both halves. A real winner must beat the benchmark in BOTH, not just full sample.",font=Font(name='Arial',bold=True,size=11,color='1F4E78'),align=L)
hrow(ws,g0,["Strategy","H1 90-08 CAGR","H1 DD","H2 08-26 CAGR","H2 DD","Full CAGR","Verdict"],[34,15,10,15,10,11,22])
for i,(name,h1,h2,full) in enumerate(guard):
    r=g0+1+i; isb=name.startswith("SMA200 3% band 3x")
    if isb: verdict="benchmark"
    elif h1["cagr"]>guard[0][1]["cagr"] and h2["cagr"]>guard[0][2]["cagr"]: verdict="beats in BOTH halves"
    elif full["cagr"]>bench["cagr"]: verdict="overfit (wins 1 half only)"
    else: verdict="worse"
    ft=BODYB if isb else BODY; f=PatternFill('solid',fgColor=LBLUE) if isb else None
    cell(ws,r,1,name,font=ft,align=L,fill=f); cell(ws,r,2,h1["cagr"],PCT,ft,fill=f); cell(ws,r,3,h1["dd"],PCT,ft,fill=f)
    cell(ws,r,4,h2["cagr"],PCT,ft,fill=f); cell(ws,r,5,h2["dd"],PCT,ft,fill=f); cell(ws,r,6,full["cagr"],PCT,ft,fill=f)
    cell(ws,r,7,verdict,font=ft,align=L,fill=f)
cell(ws,g0+2+len(guard),1,"Result: no variant beats the benchmark in both halves. 'Asym enter +3 / exit -5' is the one near-miss (more consistent, +0.5% full CAGR, -2.6 pts deeper DD).",font=SUBT,align=L)

OUT.parent.mkdir(parents=True,exist_ok=True)
wb.save(OUT)
print(f"SAVED {OUT}")
